from datetime import datetime
from io import BytesIO, StringIO
from unittest.mock import patch

from django.contrib.auth.models import Permission
from django.contrib.messages.storage.fallback import FallbackStorage
from django.contrib.sessions.middleware import SessionMiddleware
from django.core.management import call_command
from django.core.files.uploadedfile import SimpleUploadedFile
from django.test import RequestFactory
from django.test import TestCase
from django.urls import reverse
from django.utils import timezone
from openpyxl import load_workbook

from info.info_viwes.condominium.apartment.apartment_view import AUTO_DEPARTURE_NOTE, _close_active_visitants_by_plate, add_visitant, add_visitant_security, visitant_arrival, visitant_departure
from info.info_viwes.condominium.pedestrian.pedestrian_view import add_pedrestrian
from info.info_viwes.condominium.report.report_view import pedestrian_report, vehicle_report, visitant_report
from info.info_viwes.condominium.vehicle.vehicle_view import AUTO_VEHICLE_CHECKOUT_NOTE, add_vehicle, vehicle_move, vehicle_checkout_plate
from info.forms import ViewVehicleForm
from info.models import Apartment, Block, CondominiumProfile, HowTo, Notification, Pedestrian, PushNotificationToken, Resident, ResidentFeatures, Vehicle, Visitant, VisitantReport, VisitantRequiredFields
from info.utils import add_manager_notification


def _aware_datetime(value):
	return timezone.make_aware(datetime.strptime(value, "%Y-%m-%d %H:%M:%S"), timezone.get_current_timezone())


class AddManagerNotificationTests(TestCase):
	def test_employee_notifications_include_title_and_url(self):
		condominium = CondominiumProfile.objects.create_user(
			email="condominio@example.com",
			condominium_name="Condomínio Teste",
			is_active=True,
		)
		employee = CondominiumProfile.objects.create_user(
			email="porteiro@example.com",
			condominium_name="Porteiro",
			work_for=condominium,
			is_active=True,
		)
		change_visitant_permission = Permission.objects.get(codename="change_visitant")
		employee.user_permissions.add(change_visitant_permission)

		add_manager_notification(condominium, "Nova liberação de veículo", url="/condominium-visitants/")

		condo_notification = Notification.objects.get(receiver=condominium)
		employee_notification = Notification.objects.get(receiver=employee)

		self.assertEqual(condo_notification.title, "Notificação de Condomínio Teste")
		self.assertEqual(employee_notification.title, "Notificação de Condomínio Teste")
		self.assertEqual(employee_notification.url, "/condominium-visitants/")
		self.assertEqual(employee_notification.message, "Nova liberação de veículo")


class DashboardPushNotificationTests(TestCase):
	def setUp(self):
		self.user = CondominiumProfile.objects.create_user(
			email="dashboard@example.com",
			condominium_name="Dashboard Teste",
			is_active=True,
		)
		Notification.objects.create(
			condominium=self.user,
			receiver=self.user,
			title="Título",
			message="Mensagem",
			url="/alguma-url/",
		)
		PushNotificationToken.objects.create(
			user=self.user,
			token="token-teste",
			device_type="web",
		)
		for name in [
			"Dashboard > Cadastros",
			"Dashboard > Gestores",
			"Dashboard > Gestores com Moradores",
			"Dashboard > Portaria com Moradores",
		]:
			HowTo.objects.create(name=name, kind="Texto", value="ajuda")

	@patch("info.info_viwes.condominium.firebase.firebase_view.messaging.send_each_for_multicast")
	def test_dashboard_still_renders_when_push_delivery_fails(self, mocked_send):
		mocked_send.side_effect = Exception("RemoteDisconnected")
		self.client.force_login(self.user)

		response = self.client.get(reverse("info:dashboard"))

		self.assertEqual(response.status_code, 200)


class VehicleAutoCheckoutTests(TestCase):
	def setUp(self):
		self.factory = RequestFactory()
		self.condominium = CondominiumProfile.objects.create_user(
			email="portaria@example.com",
			condominium_name="Portaria Teste",
			is_active=True,
		)
		self.block = Block.objects.create(condominium=self.condominium, name="A")
		self.apartment = Apartment.objects.create(block=self.block, number=101, complement="")

	def _prepare_request(self, request):
		middleware = SessionMiddleware(lambda req: None)
		middleware.process_request(request)
		request.session.save()
		setattr(request, '_messages', FallbackStorage(request))
		request.user = self.condominium
		return request

	def test_add_vehicle_closes_previous_active_entries_with_same_plate(self):
		previous_vehicle = Vehicle.objects.create(
			condominium=self.condominium,
			protocol="OLD123",
			name="Motorista Antigo",
			document="11111111111",
			document_file=SimpleUploadedFile("old.txt", b"old vehicle", content_type="text/plain"),
			destination="A/101",
			authorized_by="Portaria",
			obs="Registro anterior",
			vehicle="Caminhão",
			vehicle_plate="ABC1234",
			arrived=True,
			has_leaved=False,
		)

		request = self._prepare_request(self.factory.post(
			"/add-vehicle",
			data={
				"name": "Motorista Novo",
				"document": "22222222222",
				"document_file": SimpleUploadedFile("new.txt", b"new vehicle", content_type="text/plain"),
				"block": str(self.block.pk),
				"apartment": str(self.apartment.pk),
				"authorized_by": "Portaria",
				"obs": "Nova entrada",
				"vehicle": "Carreta",
				"vehicle_plate": "ABC-1234",
			}
		))

		response = add_vehicle(request)

		previous_vehicle.refresh_from_db()
		new_vehicle = Vehicle.objects.exclude(pk=previous_vehicle.pk).get()

		self.assertEqual(response.status_code, 302)
		self.assertTrue(previous_vehicle.has_leaved)
		self.assertFalse(previous_vehicle.arrived)
		self.assertIn(AUTO_VEHICLE_CHECKOUT_NOTE, previous_vehicle.obs)
		self.assertEqual(new_vehicle.vehicle_plate, "ABC1234")
		self.assertFalse(new_vehicle.has_leaved)
		self.assertTrue(new_vehicle.arrived)

	def test_vehicle_move_arrival_closes_other_active_entries_with_same_plate(self):
		active_vehicle = Vehicle.objects.create(
			condominium=self.condominium,
			protocol="ACT123",
			name="Veículo Ativo",
			document="11111111111",
			document_file=SimpleUploadedFile("active.txt", b"active vehicle", content_type="text/plain"),
			destination="A/101",
			authorized_by="Portaria",
			obs="Dentro",
			vehicle="Van",
			vehicle_plate="XYZ9999",
			arrived=True,
			has_leaved=False,
		)
		returning_vehicle = Vehicle.objects.create(
			condominium=self.condominium,
			protocol="RET123",
			name="Veículo Retornando",
			document="22222222222",
			document_file=SimpleUploadedFile("return.txt", b"return vehicle", content_type="text/plain"),
			destination="A/101",
			authorized_by="Portaria",
			obs="Voltando",
			vehicle="Van",
			vehicle_plate="XYZ9999",
			arrived=False,
			has_leaved=True,
		)

		request = self._prepare_request(self.factory.get(f"/vehicle-move/{returning_vehicle.pk}"))

		response = vehicle_move(request, returning_vehicle.pk)

		active_vehicle.refresh_from_db()
		returning_vehicle.refresh_from_db()

		self.assertEqual(response.status_code, 302)
		self.assertTrue(active_vehicle.has_leaved)
		self.assertFalse(active_vehicle.arrived)
		self.assertIn(AUTO_VEHICLE_CHECKOUT_NOTE, active_vehicle.obs)
		self.assertFalse(returning_vehicle.has_leaved)
		self.assertTrue(returning_vehicle.arrived)

	def test_vehicle_checkout_plate_closes_all_active_entries_with_same_plate(self):
		first_vehicle = Vehicle.objects.create(
			condominium=self.condominium,
			protocol="BAT001",
			name="Primeiro Veículo",
			document="11111111111",
			document_file=SimpleUploadedFile("first.txt", b"first vehicle", content_type="text/plain"),
			destination="A/101",
			authorized_by="Portaria",
			obs="Primeira liberação",
			vehicle="Van",
			vehicle_plate="AAA1234",
			arrived=True,
			has_leaved=False,
		)
		second_vehicle = Vehicle.objects.create(
			condominium=self.condominium,
			protocol="BAT002",
			name="Segundo Veículo",
			document="22222222222",
			document_file=SimpleUploadedFile("second.txt", b"second vehicle", content_type="text/plain"),
			destination="A/101",
			authorized_by="Portaria",
			obs="Segunda liberação",
			vehicle="Van",
			vehicle_plate="AAA-1234",
			arrived=True,
			has_leaved=False,
		)
		other_vehicle = Vehicle.objects.create(
			condominium=self.condominium,
			protocol="BAT003",
			name="Terceiro Veículo",
			document="33333333333",
			document_file=SimpleUploadedFile("third.txt", b"third vehicle", content_type="text/plain"),
			destination="A/101",
			authorized_by="Portaria",
			obs="Outra placa",
			vehicle="Van",
			vehicle_plate="BBB1234",
			arrived=True,
			has_leaved=False,
		)

		request = self._prepare_request(self.factory.get(f"/vehicle-checkout-plate/{first_vehicle.pk}"))

		response = vehicle_checkout_plate(request, first_vehicle.pk)

		first_vehicle.refresh_from_db()
		second_vehicle.refresh_from_db()
		other_vehicle.refresh_from_db()

		self.assertEqual(response.status_code, 302)
		self.assertTrue(first_vehicle.has_leaved)
		self.assertFalse(first_vehicle.arrived)
		self.assertIn(AUTO_VEHICLE_CHECKOUT_NOTE, first_vehicle.obs)
		self.assertTrue(second_vehicle.has_leaved)
		self.assertFalse(second_vehicle.arrived)
		self.assertIn(AUTO_VEHICLE_CHECKOUT_NOTE, second_vehicle.obs)
		self.assertFalse(other_vehicle.has_leaved)
		self.assertTrue(other_vehicle.arrived)


class VisitantAutoCheckoutTests(TestCase):
	def setUp(self):
		self.factory = RequestFactory()
		self.condominium = CondominiumProfile.objects.create_user(
			email="visitantes@example.com",
			condominium_name="Visitantes Teste",
			is_active=True,
		)
		self.block = Block.objects.create(condominium=self.condominium, name="B")
		self.apartment = Apartment.objects.create(block=self.block, number=202, complement="")
		self.resident_user = CondominiumProfile.objects.create_user(
			email="morador@example.com",
			condominium_name="Morador Teste",
			resident_in=self.condominium,
			is_active=True,
		)
		self.resident = None

	def _prepare_request(self, request, user=None):
		middleware = SessionMiddleware(lambda req: None)
		middleware.process_request(request)
		request.session.save()
		setattr(request, '_messages', FallbackStorage(request))
		request.user = user or self.condominium
		return request

	def test_close_active_visitants_by_plate_adds_automatic_checkout_note(self):
		active_one = Visitant.objects.create(
			condominium=self.condominium,
			block="B",
			apartment="202",
			name="Visitante 1",
			until=_aware_datetime("2026-03-28 10:00:00"),
			vehicle_plate="AAA1234",
			arrived=True,
			visit_in=_aware_datetime("2026-03-27 09:00:00"),
		)
		active_two = Visitant.objects.create(
			condominium=self.condominium,
			block="B",
			apartment="202",
			name="Visitante 2",
			until=_aware_datetime("2026-03-28 10:00:00"),
			vehicle_plate="AAA1234",
			arrived=True,
			visit_in=_aware_datetime("2026-03-27 09:30:00"),
			comment="Observação antiga",
		)

		closed_count = _close_active_visitants_by_plate(self.condominium, "AAA-1234")

		active_one.refresh_from_db()
		active_two.refresh_from_db()

		self.assertEqual(closed_count, 2)
		self.assertFalse(active_one.arrived)
		self.assertIsNotNone(active_one.leaves_in)
		self.assertIn(AUTO_DEPARTURE_NOTE, active_one.comment)
		self.assertIn(AUTO_DEPARTURE_NOTE, active_two.comment)
		self.assertEqual(VisitantReport.objects.filter(vehicle_plate="AAA1234").count(), 2)

	def test_visitant_arrival_closes_previous_active_visitant_with_same_plate(self):
		VisitantRequiredFields.objects.create(
			condominium=self.condominium,
			photo=False,
		)
		active_visitant = Visitant.objects.create(
			condominium=self.condominium,
			block="B",
			apartment="202",
			name="Visitante Ativo",
			until=_aware_datetime("2026-03-28 10:00:00"),
			vehicle_plate="BBB1234",
			arrived=True,
			visit_in=_aware_datetime("2026-03-27 08:00:00"),
		)
		pending_visitant = Visitant.objects.create(
			condominium=self.condominium,
			block="B",
			apartment="202",
			name="Visitante Novo",
			until=_aware_datetime("2026-03-28 10:00:00"),
			vehicle_plate="BBB1234",
			arrived=False,
		)

		request = self._prepare_request(self.factory.post(
			f"/visitant-arrival/{pending_visitant.pk}",
			data={
				"block": pending_visitant.block,
				"apartment": pending_visitant.apartment,
				"name": pending_visitant.name,
				"until": "2026-03-28 10:00:00",
				"comment": "",
				"document": "123456789",
				"vehicle_model": "Sedan",
				"vehicle_plate": "BBB-1234",
			}
		))

		response = visitant_arrival(request, pending_visitant.pk)

		active_visitant.refresh_from_db()
		pending_visitant.refresh_from_db()

		self.assertEqual(response.status_code, 302)
		self.assertFalse(active_visitant.arrived)
		self.assertIsNotNone(active_visitant.leaves_in)
		self.assertIn(AUTO_DEPARTURE_NOTE, active_visitant.comment)
		self.assertTrue(pending_visitant.arrived)
		self.assertEqual(pending_visitant.vehicle_plate, "BBB1234")

	def test_visitant_arrival_invalid_form_does_not_mark_arrived(self):
		VisitantRequiredFields.objects.create(
			condominium=self.condominium,
			photo=False,
		)
		pending_visitant = Visitant.objects.create(
			condominium=self.condominium,
			block="B",
			apartment="202",
			name="Visitante Inválido",
			until=_aware_datetime("2026-03-28 10:00:00"),
			arrived=False,
		)

		request = self._prepare_request(self.factory.post(
			f"/visitant-arrival/{pending_visitant.pk}",
			data={
				"block": pending_visitant.block,
				"apartment": pending_visitant.apartment,
				"name": pending_visitant.name,
				"until": "2026-03-28 10:00:00",
				"comment": "",
				"vehicle_model": "Sedan",
			}
		))

		response = visitant_arrival(request, pending_visitant.pk)

		pending_visitant.refresh_from_db()

		self.assertEqual(response.status_code, 200)
		self.assertFalse(pending_visitant.arrived)
		self.assertIsNone(pending_visitant.visit_in)

	def test_visitant_departure_closes_all_active_entries_with_same_plate(self):
		first_visitant = Visitant.objects.create(
			condominium=self.condominium,
			block="B",
			apartment="202",
			name="Visitante 1",
			until=_aware_datetime("2026-03-28 10:00:00"),
			vehicle_plate="CCC1234",
			arrived=True,
			can_leave=True,
			visit_in=_aware_datetime("2026-03-27 08:00:00"),
		)
		second_visitant = Visitant.objects.create(
			condominium=self.condominium,
			block="B",
			apartment="202",
			name="Visitante 2",
			until=_aware_datetime("2026-03-28 10:00:00"),
			vehicle_plate="CCC-1234",
			arrived=True,
			can_leave=True,
			visit_in=_aware_datetime("2026-03-27 08:30:00"),
			comment="Observação",
		)
		other_visitant = Visitant.objects.create(
			condominium=self.condominium,
			block="B",
			apartment="202",
			name="Visitante 3",
			until=_aware_datetime("2026-03-28 10:00:00"),
			vehicle_plate="DDD1234",
			arrived=True,
			can_leave=True,
			visit_in=_aware_datetime("2026-03-27 09:00:00"),
		)

		request = self._prepare_request(self.factory.get(f"/visitant-departure/{first_visitant.pk}"))

		response = visitant_departure(request, first_visitant.pk)

		first_visitant.refresh_from_db()
		second_visitant.refresh_from_db()
		other_visitant.refresh_from_db()

		self.assertEqual(response.status_code, 302)
		self.assertFalse(first_visitant.arrived)
		self.assertIsNotNone(first_visitant.leaves_in)
		self.assertIn(AUTO_DEPARTURE_NOTE, first_visitant.comment)
		self.assertFalse(second_visitant.arrived)
		self.assertIsNotNone(second_visitant.leaves_in)
		self.assertIn(AUTO_DEPARTURE_NOTE, second_visitant.comment)
		self.assertTrue(other_visitant.arrived)
		self.assertIsNone(other_visitant.leaves_in)
		self.assertEqual(VisitantReport.objects.filter(vehicle_plate="CCC1234").count(), 1)
		self.assertEqual(VisitantReport.objects.filter(vehicle_plate="CCC-1234").count(), 1)

	def test_add_visitant_security_invalid_form_does_not_crash(self):
		VisitantRequiredFields.objects.create(
			condominium=self.condominium,
			photo=False,
		)

		request = self._prepare_request(self.factory.post(
			"/add-visitant-security",
			data={
				"name": "Visitante sem apartamento",
				"block": str(self.block.pk),
				"until": "2026-03-28T10:00",
				"document": "123456789",
			}
		))

		response = add_visitant_security(request)

		self.assertEqual(response.status_code, 200)
		self.assertEqual(Visitant.objects.count(), 0)


class RoutingRegressionTests(TestCase):
	def test_named_routes_reverse_to_expected_paths(self):
		self.assertEqual(reverse("info:delete-file", args=[15]), "/delete-file/15")
		self.assertEqual(reverse("info:visitants-aux"), "/visitants-aux")
		self.assertEqual(reverse("info:remove-permanent-visitants"), "/visitants-aux/")

	def test_delete_condominium_redirects_back_to_list(self):
		request_factory = RequestFactory()
		admin = CondominiumProfile.objects.create_user(
			email="admin@example.com",
			condominium_name="Administrador",
			is_active=True,
		)
		condominium = CondominiumProfile.objects.create_user(
			email="condo-delete@example.com",
			condominium_name="Condomínio para apagar",
			is_active=True,
		)
		request = request_factory.get(f"/admin-delete-condominium/{condominium.pk}")
		request.user = admin

		from info.info_viwes.administrator_view import delete_condominium

		response = delete_condominium(request, condominium.pk)

		self.assertEqual(response.status_code, 302)
		self.assertEqual(response.url, reverse("info:admin-condominiums"))
		self.assertFalse(CondominiumProfile.objects.filter(pk=condominium.pk).exists())


class AccessReleaseFormValidationTests(TestCase):
	def setUp(self):
		self.factory = RequestFactory()
		self.condominium = CondominiumProfile.objects.create_user(
			email="acesso@example.com",
			condominium_name="Acesso Teste",
			is_active=True,
		)
		self.block = Block.objects.create(condominium=self.condominium, name="C")
		self.apartment = Apartment.objects.create(block=self.block, number=303, complement="")

	def _prepare_request(self, request):
		middleware = SessionMiddleware(lambda req: None)
		middleware.process_request(request)
		request.session.save()
		setattr(request, '_messages', FallbackStorage(request))
		request.user = self.condominium
		return request

	def test_add_vehicle_invalid_form_does_not_crash(self):
		request = self._prepare_request(self.factory.post(
			"/add-vehicle",
			data={
				"name": "Veículo sem apartamento",
				"document": "123456",
				"block": str(self.block.pk),
				"authorized_by": "Portaria",
				"obs": "Teste",
			}
		))

		response = add_vehicle(request)

		self.assertEqual(response.status_code, 200)
		self.assertEqual(Vehicle.objects.count(), 0)

	def test_add_pedestrian_invalid_form_does_not_crash(self):
		request = self._prepare_request(self.factory.post(
			"/add-pedestrian",
			data={
				"name": "Pedestre sem apartamento",
				"document": "654321",
				"block": str(self.block.pk),
				"authorized_by": "Portaria",
				"obs": "Teste",
			}
		))

		response = add_pedrestrian(request)

		self.assertEqual(response.status_code, 200)
		self.assertEqual(Pedestrian.objects.count(), 0)

	def test_view_vehicle_form_accepts_vehicle_instance(self):
		vehicle = Vehicle.objects.create(
			condominium=self.condominium,
			protocol="VV123",
			name="Motorista",
			document="999",
			document_file=SimpleUploadedFile("doc.txt", b"doc", content_type="text/plain"),
			destination="C/303",
			authorized_by="Portaria",
			obs="Liberado",
			vehicle="Carro",
			vehicle_plate="AAA1234",
			arrived=True,
			has_leaved=False,
		)

		form = ViewVehicleForm(instance=vehicle)

		self.assertEqual(form.instance, vehicle)


class ResidentVisitantReleaseValidationTests(TestCase):
	def setUp(self):
		self.factory = RequestFactory()
		self.condominium = CondominiumProfile.objects.create_user(
			email="condominio-morador@example.com",
			condominium_name="Condomínio Morador",
			is_active=True,
		)
		ResidentFeatures.objects.create(condominium=self.condominium, permanent=True)
		VisitantRequiredFields.objects.create(
			condominium=self.condominium,
			photo=False,
			vehicle_plate=False,
		)
		self.block = Block.objects.create(condominium=self.condominium, name="D")
		self.apartment = Apartment.objects.create(block=self.block, number=404, complement="")
		self.resident_user = CondominiumProfile.objects.create_user(
			email="morador-liberacao@example.com",
			condominium_name="Morador Liberação",
			resident_in=self.condominium,
			is_active=True,
		)
		Resident.objects.create(
			apartment=self.apartment,
			name=self.resident_user.condominium_name,
			email=self.resident_user.email,
		)

	def _prepare_request(self, request):
		middleware = SessionMiddleware(lambda req: None)
		middleware.process_request(request)
		request.session.save()
		setattr(request, '_messages', FallbackStorage(request))
		request.user = self.resident_user
		return request

	def test_resident_release_requires_name(self):
		request = self._prepare_request(self.factory.post(
			"/add-visitants/",
			data={
				"until": "2026-03-28T10:00",
				"comment": "Teste",
			}
		))

		response = add_visitant(request)

		self.assertEqual(response.status_code, 200)
		self.assertEqual(Visitant.objects.count(), 0)

	def test_permanent_resident_release_does_not_require_until(self):
		request = self._prepare_request(self.factory.post(
			"/add-visitants/",
			data={
				"name": "Visitante Permanente",
				"permanent": "on",
				"comment": "Teste",
			}
		))

		response = add_visitant(request)

		self.assertEqual(response.status_code, 302)
		self.assertTrue(Visitant.objects.filter(name="Visitante Permanente").exists())


class SanitizeDuplicateActivePlatesCommandTests(TestCase):
	def setUp(self):
		self.condominium = CondominiumProfile.objects.create_user(
			email="sanitiza@example.com",
			condominium_name="Sanitiza Teste",
			is_active=True,
		)

	def test_command_dry_run_does_not_change_vehicle_records(self):
		older_vehicle = Vehicle.objects.create(
			condominium=self.condominium,
			protocol="OLD-A",
			name="Mais Antigo",
			document="111",
			document_file=SimpleUploadedFile("olda.txt", b"olda", content_type="text/plain"),
			destination="A/1",
			authorized_by="Portaria",
			obs="Antigo",
			vehicle="Carro",
			vehicle_plate="CCC1234",
			arrived=True,
			has_leaved=False,
		)
		Vehicle.objects.create(
			condominium=self.condominium,
			protocol="NEW-A",
			name="Mais Novo",
			document="222",
			document_file=SimpleUploadedFile("newa.txt", b"newa", content_type="text/plain"),
			destination="A/1",
			authorized_by="Portaria",
			obs="Novo",
			vehicle="Carro",
			vehicle_plate="CCC1234",
			arrived=True,
			has_leaved=False,
		)

		stdout = StringIO()
		call_command('sanitize_duplicate_active_plates', '--scope', 'vehicles', stdout=stdout)

		older_vehicle.refresh_from_db()
		self.assertFalse(older_vehicle.has_leaved)
		self.assertTrue(older_vehicle.arrived)
		self.assertIn('Modo simulação ativado', stdout.getvalue())

	def test_command_apply_closes_stale_visitant_and_creates_report(self):
		older_visitant = Visitant.objects.create(
			condominium=self.condominium,
			block="C",
			apartment="303",
			name="Visitante Antigo",
			until=_aware_datetime("2026-03-28 10:00:00"),
			vehicle_plate="DDD1234",
			arrived=True,
			visit_in=_aware_datetime("2026-03-27 08:00:00"),
		)
		Visitant.objects.create(
			condominium=self.condominium,
			block="C",
			apartment="303",
			name="Visitante Novo",
			until=_aware_datetime("2026-03-28 10:00:00"),
			vehicle_plate="DDD1234",
			arrived=True,
			visit_in=_aware_datetime("2026-03-27 09:00:00"),
		)

		stdout = StringIO()
		call_command('sanitize_duplicate_active_plates', '--scope', 'visitants', '--apply', stdout=stdout)

		older_visitant.refresh_from_db()
		self.assertFalse(older_visitant.arrived)
		self.assertIsNotNone(older_visitant.leaves_in)
		self.assertIn(AUTO_DEPARTURE_NOTE, older_visitant.comment)
		self.assertTrue(VisitantReport.objects.filter(vehicle_plate='DDD1234').exists())
		self.assertIn('Modo aplicação ativado', stdout.getvalue())


class AccessReportRegressionTests(TestCase):
	def setUp(self):
		self.factory = RequestFactory()
		self.condominium = CondominiumProfile.objects.create_user(
			email="relatorios@example.com",
			condominium_name="Relatórios Teste",
			is_active=True,
		)

	def _prepare_request(self, request):
		middleware = SessionMiddleware(lambda req: None)
		middleware.process_request(request)
		request.session.save()
		setattr(request, '_messages', FallbackStorage(request))
		request.user = self.condominium
		return request

	def test_pedestrian_xls_does_not_show_exit_for_active_entry(self):
		pedestrian = Pedestrian.objects.create(
			condominium=self.condominium,
			protocol="PED123",
			name="Pedestre Ativo",
			document="123",
			document_file=SimpleUploadedFile("ped.txt", b"ped", content_type="text/plain"),
			destination="A/101",
			authorized_by="Portaria",
			obs="Ativo",
			arrived=True,
			has_leaved=False,
		)

		request = self._prepare_request(self.factory.post(
			"/pedestrian-report",
			data={
				"visits_from": pedestrian.created.strftime("%Y-%m-%d"),
				"visits_until": pedestrian.created.strftime("%Y-%m-%d"),
				"block": "",
				"type": "2",
			}
		))

		response = pedestrian_report(request)

		workbook = load_workbook(BytesIO(response.content))
		worksheet = workbook.active

		self.assertEqual(response.status_code, 200)
		self.assertIsNone(worksheet["I4"].value)

	def test_vehicle_xls_uses_vehicle_title_and_hides_exit_for_active_entry(self):
		vehicle = Vehicle.objects.create(
			condominium=self.condominium,
			protocol="VEI123",
			name="Veículo Ativo",
			document="456",
			document_file=SimpleUploadedFile("veh.txt", b"veh", content_type="text/plain"),
			destination="B/202",
			authorized_by="Portaria",
			obs="Ativo",
			vehicle="Carro",
			vehicle_plate="ABC1234",
			arrived=True,
			has_leaved=False,
		)

		request = self._prepare_request(self.factory.post(
			"/vehicle-report",
			data={
				"visits_from": vehicle.created.strftime("%Y-%m-%d"),
				"visits_until": vehicle.created.strftime("%Y-%m-%d"),
				"block": "",
				"type": "2",
			}
		))

		response = vehicle_report(request)

		workbook = load_workbook(BytesIO(response.content))
		worksheet = workbook.active

		self.assertEqual(response.status_code, 200)
		self.assertIn("Relatório de Veículos", worksheet["A2"].value)
		self.assertIsNone(worksheet["K4"].value)

	def test_visitant_xls_handles_missing_resident_and_until(self):
		visitant = VisitantReport.objects.create(
			condominium=self.condominium,
			block="C",
			apartment="303",
			name="Visitante sem morador",
			document="789",
			security_name="Portaria",
			visit_in=timezone.now(),
			leaves_in=None,
			resident=None,
			until=None,
		)

		request = self._prepare_request(self.factory.post(
			"/visitant-report",
			data={
				"visits_from": visitant.created.strftime("%Y-%m-%d"),
				"visits_until": visitant.created.strftime("%Y-%m-%d"),
				"block": "",
				"type": "2",
			}
		))

		response = visitant_report(request)

		workbook = load_workbook(BytesIO(response.content))
		worksheet = workbook.active

		self.assertEqual(response.status_code, 200)
		self.assertIsNone(worksheet["D4"].value)
		self.assertIsNone(worksheet["E4"].value)
		self.assertIsNone(worksheet["H4"].value)
