import os
from datetime import datetime, timedelta, date
from io import BytesIO

import pytz
from django.contrib import messages
from django.contrib.auth.decorators import login_required
from django.contrib.sites.shortcuts import get_current_site
from django.db.models import Sum, Avg
from django.http import HttpResponse
from django.shortcuts import render, redirect
from django.urls import reverse
from django.utils.dateparse import parse_date
from django.utils.timezone import make_aware
from openpyxl import Workbook
from openpyxl.styles import *

from fpdf import FPDF

from info.forms import ActivityReportForm, GeneralReportForm, MessageReportForm, ChecklistReportForm, VisitantReportForm
from info.models import Informative, Function, FunctionItem, ImageModel, FunctionItemFileModel, Order, Message, \
    MessageFileModel, Checklist, Task, SurveyOptionModel, SurveyModel, FunctionItemVideoLink, Contract, UserLocation, \
    Resident, ReviewAnswer, Review, Visitant, VisitantReport, StorageEntry, Product, InformativeKind, Timeline, \
    TimelinePhase, ResidentActivity, ResidentActivityAnswer, SurveyAnswerModel, Reservation, UserControl, ReportLogo, \
    ActivityFunction, Pedestrian, Vehicle
from info.utils import get_condominium


FIXED_TZ = pytz.timezone("America/Sao_Paulo")


def _build_report_datetime_range(initial, until):
    initial_date = parse_date(initial)
    until_date = parse_date(until)

    if not initial_date or not until_date:
        return None, None

    start_datetime = make_aware(datetime.combine(initial_date, datetime.min.time()), FIXED_TZ)
    end_datetime = make_aware(datetime.combine(until_date + timedelta(days=1), datetime.min.time()), FIXED_TZ)
    return start_datetime, end_datetime


@login_required(login_url='info:sign-in')
def reports(request):
    condominium = get_condominium(request)
    context = {'user': condominium}
    return render(request, "info/condominium/report/reports.html", context=context)


@login_required(login_url='info:sign-in')
def activity_report(request):
    condominium = get_condominium(request)
    informative_kind = InformativeKind.objects.filter(condominium=condominium)
    form = ActivityReportForm(informative_kind=informative_kind)

    if request.method == "POST":

        initial = request.POST.get('initial')
        until = request.POST.get('until')

        new_until = None
        if parse_date(until) == date.today():
            new_until = date.today() + timedelta(1)
        else:
            new_until = None

        if initial == until:
            new_until = parse_date(initial) + timedelta(1)

        kind = request.POST.get('kind')
        request.session['initial'] = initial
        request.session['until'] = until
        request.session['kind'] = kind

        if new_until:
            until_filter = new_until
        else:
            until_filter = until

        if kind == "TODAS":

            objects_last_days = Informative.objects.filter(created__gte=initial, created__lte=until_filter,
                                                           condominium=condominium)
            kind = "TODAS as atividades"
        else:
            informative_kind = InformativeKind.objects.get(pk=int(kind))
            objects_last_days = Informative.objects.filter(created__gte=initial, created__lte=until_filter,
                                                           condominium=condominium,
                                                           kind__iexact=informative_kind.name)

        return _activity_pdf(request, condominium, objects_last_days, kind, initial, until)
    context = {'form': form,
               }
    return render(request, "info/condominium/report/activity_report.html", context=context)


def _activity_pdf(request, condominium, activities, kind, initial, until):
    if len(activities) > 0:

        pdf = PDF(test=condominium.is_testing)
        pdf.add_page()
        pdf.set_auto_page_break(auto=True, margin=15)

        _add_pdf_header(pdf, condominium)

        pdf.set_font("helvetica", "B", 12)
        pdf.cell(0, 10,
                 txt=f"Relatório das atividades de {datetime.strptime(initial, '%Y-%m-%d').strftime('%d/%m/%Y')} até {datetime.strptime(until, '%Y-%m-%d').strftime('%d/%m/%Y')}",
                 align="C", ln=True)
        pdf.cell(0, 12,
                 txt=f"{kind}",
                 align="C", ln=True)
        pdf.ln(15)
        pdf.set_font("helvetica", "", 10)
        pdf.cell(145, 7, txt="Atividades:", ln=False)
        pdf.cell(50, 7, txt=f"Total de atividades: {len(activities)}", ln=True)

        bullet = "-"
        nested_bullet = ">"
        for idx, obj in enumerate(activities, start=1):
            pdf.cell(15)
            pdf.cell(200, 7, txt=f"{idx}. Atividade: {obj.title}", ln=True)

            if obj.location:
                pdf.cell(25)
                pdf.cell(200, 7, txt=f"{bullet} Atividade realizada por { obj.location.condominium.condominium_name }, em { obj.location.address }", ln=True)

            pdf.cell(25)
            pdf.cell(200, 7, txt=f"{bullet} Funções:", ln=True)

            y_start = None
            functions = ActivityFunction.objects.filter(informative=obj)
            for idx_2, func in enumerate(functions, start=1):
                pdf.cell(35)
                pdf.cell(200, 7, txt=f"{nested_bullet} {func.title}:", ln=True)
                if func.description:
                    pdf.cell(35)
                    pdf.cell(200, 7, txt=f"{nested_bullet} {func.description}:", ln=True)

                files = FunctionItemFileModel.objects.filter(function_item=func)
                if files.exists():
                    for idx_3, file in enumerate(files, start=1):
                        pdf.cell(35)
                        pdf.set_text_color(0, 0, 255)
                        pdf.cell(200, 7, txt=f"{get_current_site(request).domain + file.file.url}", ln=True,
                                 link="http://" + get_current_site(request).domain + file.file.url)

                if func.link:
                    pdf.cell(35)
                    pdf.set_text_color(0, 0, 0)
                    pdf.cell(200, 7, txt=f"{nested_bullet} Vídeo:", ln=True)
                    pdf.cell(35)
                    pdf.set_text_color(0, 0, 255)
                    pdf.cell(200, 7, txt=f"{get_current_site(request).domain + func.link}", ln=True,
                             link=func.link)

                images = ImageModel.objects.filter(function_item=func)
                if images.exists():
                    pdf.cell(35)
                    pdf.set_text_color(0, 0, 0)
                    pdf.cell(200, 7, txt=f"{nested_bullet} Imagens:", ln=True)
                    y_start = pdf.get_y()
                    for idx_3, img in enumerate(images, start=1):
                        file_path = img.image.path
                        if os.path.exists(file_path):
                            pdf.image(img.image.path, pdf.w/2 - 30, y_start, 50, 50)
                            y_start = y_start + 60
                else:
                    y_start = pdf.get_y()

            pdf.set_text_color(0, 0, 0)

            pdf.ln(7)
            line_start = y_start if y_start else pdf.get_y()
            pdf.line(25, line_start, 185, line_start)
            pdf.ln(15)

        messages.success(request, "Relatório Gerado!")
        return _make_pdf_response(pdf, "Relatório_Atividades")

    messages.success(request, "Nenhuma atividade realizada neste período!")
    return redirect(reverse('info:reports'))


@login_required(login_url='info:sign-in')
def order_report(request):
    condominium = get_condominium(request)
    form = GeneralReportForm()

    if request.method == "POST":
        initial = request.POST.get('initial')
        until = request.POST.get('until')
        request.session['initial'] = initial
        request.session['until'] = until

        if parse_date(until) == date.today():
            until_filter = date.today() + timedelta(1)
        else:
            until_filter = until

        if initial == until:
            until_filter = parse_date(initial) + timedelta(1)

        objects_last_days = Order.objects.filter(created__gte=initial, created__lte=until_filter,
                                                 apartment__block__condominium=condominium)

        return _order_pdf(request, condominium, objects_last_days, initial, until)
    context = {'form': form,
               }
    return render(request, "info/condominium/report/order_report.html", context=context)


@login_required(login_url='info:sign-in')
def export_order_to_pdf(request, id):
    condominium = get_condominium(request)

    order = [Order.objects.get(pk=id)]

    return _order_pdf(request, condominium, order, 0, False)


def _order_pdf(request, condominium, orders, initial, until):
    if len(orders) > 0:
        context = {'condominium': condominium,
                   'orders': orders,
                   'test': condominium.is_testing}

        if initial != 0:
            context['initial'] = parse_date(initial).strftime("%d/%m/%y")

        if until:
            context['until'] = parse_date(until).strftime("%d/%m/%y")


        pdf = PDF(test=condominium.is_testing)
        pdf.add_page()
        pdf.set_auto_page_break(auto=True, margin=15)

        _add_pdf_header(pdf, condominium)

        pdf.set_font("helvetica", "B", 12)
        if initial != 0:
            file_name = "Relatório_Correspondencias"
            pdf.cell(0, 10,
                     txt=f"Relatório das correspondências de {datetime.strptime(initial, '%Y-%m-%d').strftime('%d/%m/%Y')} até {datetime.strptime(until, '%Y-%m-%d').strftime('%d/%m/%Y')}",
                     align="C", ln=True)
        else:
            file_name = "Correspondencia_" + orders[0].code
            pdf.cell(0, 10,
                     txt=f"Correspondência",
                     align="C", ln=True)

        pdf.ln(15)
        pdf.set_font("helvetica", "", 10)
        pdf.cell(145, 7, txt="Correspondências:", ln=False)
        if initial != 0:
            pdf.cell(50, 7, txt=f"Total de atividades: {len(orders)}", ln=True)

        y_start = None
        bullet = "-"
        for idx, obj in enumerate(orders, start=1):
            pdf.cell(15)
            pdf.cell(200, 7, txt=f"{idx}. Nome: {obj.name}", ln=True)

            if obj.description:
                pdf.cell(25)
                pdf.cell(200, 7, txt=f"{bullet} Descrição: {obj.description}", ln=True)

            if obj.received_by:
                pdf.cell(25)
                pdf.cell(200, 7, txt=f"{bullet} Recebida por: {obj.received_by}", ln=True)

            if obj.apartment:
                pdf.cell(25)
                pdf.cell(200, 7, txt=f"{bullet} Apartamento: { obj.apartment.number } { obj.apartment.complement }", ln=True)

            pdf.cell(25)
            pdf.cell(200, 7, txt=f"{bullet} Código: {obj.code}", ln=True)
            pdf.cell(25)
            pdf.cell(200, 7, txt=f"{bullet} Morador notificado em: {obj.created.astimezone(FIXED_TZ).strftime('%d/%m/%y')}", ln=True)

            if obj.delivered:
                pdf.cell(25)
                pdf.cell(200, 7, txt=f"{bullet} Entregue em: {obj.delivered.astimezone(FIXED_TZ).strftime('%d/%m/%y')}", ln=True)
                pdf.cell(25)
                pdf.cell(200, 7, txt=f"{bullet} Entregue por: { obj.delivered_by }", ln=True)
                pdf.cell(25)
                pdf.cell(200, 7, txt=f"{bullet} Retirada por: { obj.collected_by }", ln=True)
            else:
                pdf.cell(25)
                pdf.cell(200, 7, txt=f"{bullet} Ainda não retirada pelo morador", ln=True)

            pdf.ln(15)
            if obj.image:
                y_start = pdf.get_y()
                pdf.image(obj.image.path, pdf.w / 2 - 30, y_start, 50, 50)
                y_start = y_start + 60

            pdf.ln(7)
            line_start = y_start if y_start else pdf.get_y()
            pdf.line(25, line_start, 185, line_start)
            pdf.ln(15)

        messages.success(request, "Documento Gerado!")
        return _make_pdf_response(pdf, file_name)

    messages.success(request, "Nenhuma correspondência recebida neste período!")
    return redirect(reverse('info:reports'))


@login_required(login_url='info:sign-in')
def message_report(request):
    condominium = get_condominium(request)
    form = MessageReportForm()

    if request.method == "POST":

        initial = request.POST.get('initial')
        until = request.POST.get('until')
        kind = request.POST.get('kind')

        request.session['initial'] = initial
        request.session['until'] = until
        request.session['kind'] = kind

        if parse_date(until) == date.today():
            until_filter = date.today() + timedelta(1)
        else:
            until_filter = until

        if initial == until:
            until_filter = parse_date(initial) + timedelta(1)

        if kind == "TODOS":
            messages_last_days = Message.objects.filter(created__gte=initial, created__lte=until_filter,
                                                        condominium=condominium)
            kind = ""
        else:
            messages_last_days = Message.objects.filter(created__gte=initial, created__lte=until_filter,
                                                        condominium=condominium,
                                                        kind__iexact=kind)

        return _message_pdf(request, condominium, messages_last_days, kind, initial, until)
    context = {'form': form,
               }
    return render(request, "info/condominium/report/message_report.html", context=context)


@login_required(login_url='info:sign-in')
def export_message_to_pdf(request, id):
    condominium = get_condominium(request)

    message = [Message.objects.get(pk=id)]

    return _message_pdf(request, condominium, message, message[0].kind, 0, 0)


def _message_pdf(request, condominium, messages_list, kind, initial, until):
    if len(messages_list) > 0:

        pdf = PDF(test=condominium.is_testing)
        pdf.add_page()
        pdf.set_auto_page_break(auto=True, margin=15)

        _add_pdf_header(pdf, condominium)

        pdf.set_font("helvetica", "B", 12)
        if initial != 0:
            file_name = "Relatório_Mensagens"
            pdf.cell(0, 10,
                     txt=f"Relatório das mensagens de {datetime.strptime(initial, '%Y-%m-%d').strftime('%d/%m/%Y')} até {datetime.strptime(until, '%Y-%m-%d').strftime('%d/%m/%Y')}",
                     align="C", ln=True)
        else:
            file_name = "Mensagem_" + str(messages_list[0].pk)
            pdf.cell(0, 10,
                     txt=f"Mensagem",
                     align="C", ln=True)

        pdf.ln(15)
        y_start = None
        bullet = "-"
        for idx, message in enumerate(messages_list, start=1):
            files = MessageFileModel.objects.filter(message=message)
            pdf.cell(15)
            pdf.cell(200, 7, txt=f"{idx}. Tipo: {message.kind}", ln=True)

            if message.block:
                pdf.cell(15)
                pdf.cell(200, 7, txt=f"{bullet}. Para o bloco: {message.block}", ln=True)

            if message.apartment:
                pdf.cell(15)
                pdf.cell(200, 7, txt=f"{bullet}. Para o apartamento: {message.apartment}", ln=True)

            if message.message:
                pdf.cell(15)
                pdf.cell(200, 7, txt=f"{bullet}. Mensagem: {message.message}", ln=True)

            pdf.cell(15)
            pdf.cell(200, 7, txt=f"{bullet}. Enviada em: {message.created.astimezone(FIXED_TZ).strftime('%d/%m/%y - %H:%M:%S')}", ln=True)

            if len(files) > 0:
                pdf.cell(15)
                pdf.cell(200, 7, txt=f"{bullet}. Anexos:", ln=True)
                for idx_2, file in enumerate(files, start=1):
                    pdf.cell(35)
                    pdf.set_text_color(0, 0, 255)
                    pdf.cell(200, 7, txt=f"{get_current_site(request).domain + file.file.url}", ln=True,
                             link="http://" + get_current_site(request).domain + file.file.url)

            pdf.ln(7)
            line_start = y_start if y_start else pdf.get_y()
            pdf.line(25, line_start, 185, line_start)
            pdf.ln(15)

        messages.success(request, "Documento Gerado!")
        return _make_pdf_response(pdf, file_name)
    messages.success(request, "Nenhum comunicado enviado neste período!")
    return redirect(reverse('info:reports'))


@login_required(login_url='info:sign-in')
def checklist_report(request):
    condominium = get_condominium(request)
    form = ChecklistReportForm()

    if request.method == "POST":
        initial = request.POST.get('initial')
        until = request.POST.get('until')
        kind = request.POST.get('kind')

        request.session['initial'] = initial
        request.session['until'] = until
        request.session['kind'] = kind

        if parse_date(until) == date.today():
            until_filter = date.today() + timedelta(1)
        else:
            until_filter = until

        if initial == until:
            until_filter = parse_date(initial) + timedelta(1)

        checklists_last_days = Checklist.objects.filter(created__gte=initial, created__lte=until_filter,
                                                        condominium=condominium)

        return _checklist_pdf(request, condominium, checklists_last_days, kind, initial, until)

    context = {'form': form,
               }
    return render(request, "info/condominium/report/checklist_report.html", context=context)


def _checklist_pdf(request, condominium, checklists, kind, initial, until):
    if len(checklists) > 0:

        pdf = PDF(test=condominium.is_testing)
        pdf.add_page()
        pdf.set_auto_page_break(auto=True, margin=15)

        _add_pdf_header(pdf, condominium)

        pdf.set_font("helvetica", "B", 12)
        if initial != 0:
            file_name = "Relatório_Checklists"
            pdf.cell(0, 10,
                     txt=f"Relatório dos Checklists de {datetime.strptime(initial, '%Y-%m-%d').strftime('%d/%m/%Y')} até {datetime.strptime(until, '%Y-%m-%d').strftime('%d/%m/%Y')}",
                     align="C", ln=True)
        else:
            file_name = "Cecklist_" + str(checklists[0].pk)
            pdf.cell(0, 10,
                     txt=f"Checklist",
                     align="C", ln=True)

        pdf.ln(15)
        bullet = "-"
        nested_bullet = ">"

        for idx, checklist in enumerate(checklists, start=1):
            checklist_tasks = Task.objects.filter(checklist=checklist)
            pdf.cell(15)
            pdf.cell(200, 7, txt=f"{idx}. Nome: {checklist.title}", ln=True)
            pdf.cell(15)
            pdf.cell(200, 7, txt=f"{bullet}. Criado em: {checklist.created.astimezone(FIXED_TZ).strftime('%d/%m/%y - %H:%M:%S')}", ln=True)
            pdf.cell(15)
            pdf.cell(200, 7, txt=f"{bullet}. Verificações:", ln=True)
            for idx_2, task in enumerate(checklist_tasks, start=1):
                pdf.cell(35)
                pdf.cell(200, 7, txt=f"{bullet}. Checkagem: {task.task_name}", ln=True)
                if task.is_completed:
                    pdf.cell(35)
                    pdf.cell(200, 7, txt=f"{bullet}. Situação: CONCLUÍDO", ln=True)
                elif task.reported_problem:
                    pdf.cell(35)
                    pdf.cell(200, 7, txt=f"{bullet}. Situação: COM PROBLEMA", ln=True)
                    pdf.cell(45)
                    pdf.cell(200, 7, txt=f"{nested_bullet}. Descrição: {task.problem_description}", ln=True)
                    if task.reported_problem_image:
                        y_start = pdf.get_y()
                        pdf.image(task.reported_problem_image.path, pdf.w / 2 - 30, y_start, 50, 50)
                        pdf.ln(60)
                else:
                    pdf.cell(35)
                    pdf.cell(200, 7, txt=f"{bullet}. Situação: FALTA VERIFICAÇÃO", ln=True)

                if task.location:
                    pdf.cell(15)
                    pdf.cell(200, 7, txt=f"{bullet}. Verificado por: {task.location.condominium.condominium_name} em {task.location.address}", ln=True)

            if checklist.location:
                pdf.cell(15)
                pdf.cell(200, 7, txt=f"{bullet}. Checklist criado por: { checklist.location.condominium.condominium_name } em {checklist.location.address}", ln=True)

            pdf.ln(7)
            line_start = pdf.get_y()
            pdf.line(25, line_start, 185, line_start)
            pdf.ln(15)

        messages.success(request, "Documento Gerado!")
        return _make_pdf_response(pdf, file_name)
    messages.success(request, "Nenhum checklist com esta situação neste período!")
    return redirect(reverse('info:reports'))


@login_required(login_url='info:sign-in')
def view_checklist_report(request, id):
    condominium = get_condominium(request)

    kind = "TODAS"
    checklists = [Checklist.objects.get(pk=id)]
    return _checklist_pdf(request, condominium, checklists, kind, 0, 0)


@login_required(login_url='info:sign-in')
def surveys_report(request):
    condominium = get_condominium(request)
    form = GeneralReportForm()

    if request.method == "POST":
        initial = request.POST.get('initial')
        until = request.POST.get('until')

        request.session['initial'] = initial
        request.session['until'] = until

        if parse_date(until) == date.today():
            until_filter = date.today() + timedelta(1)
        else:
            until_filter = until

        if initial == until:
            until_filter = parse_date(initial) + timedelta(1)

        objects_last_days = SurveyModel.objects.filter(created__gte=initial, created__lte=until_filter,
                                                       condominium=condominium)

        return _survey_pdf(request, condominium, objects_last_days, initial, until, False)
    context = {'form': form,
               }
    return render(request, "info/condominium/report/order_report.html", context=context)


@login_required(login_url='info:sign-in')
def survey_report(request, id):
    condominium = get_condominium(request)
    survey = [SurveyModel.objects.get(pk=id)]

    return _survey_pdf(request, condominium, survey, 0, 0, True)


def _survey_pdf(request, condominium, surveys, initial, until, end, answer=None):
    if len(surveys) > 0:

        pdf = PDF(test=condominium.is_testing)
        pdf.add_page()
        pdf.set_auto_page_break(auto=True, margin=15)

        _add_pdf_header(pdf, condominium)

        pdf.set_font("helvetica", "B", 12)
        if initial != 0:
            file_name = "Relatório_Avaliações"
            pdf.cell(0, 10,
                     txt=f"Relatório das enquetes de {datetime.strptime(initial, '%Y-%m-%d').strftime('%d/%m/%Y')} até {datetime.strptime(until, '%Y-%m-%d').strftime('%d/%m/%Y')}",
                     align="C", ln=True)
        else:
            file_name = "Enquete_" + str(surveys[0].pk)

        pdf.ln(15)
        bullet = "-"
        nested_bullet = ">"

        for idx, survey in enumerate(surveys, start=1):
            options = SurveyOptionModel.objects.filter(survey=survey)
            ansers = SurveyAnswerModel.objects.filter(survey=survey)
            pdf.cell(15)
            pdf.cell(200, 7, txt=f"{idx}. ENQUETE:", ln=True)
            pdf.cell(15)
            pdf.cell(200, 7, txt=f"{bullet}. Pergunta:", ln=True)
            pdf.cell(25)
            pdf.cell(200, 7, txt=f"{nested_bullet} {survey.question}", ln=True)
            pdf.cell(15)
            pdf.cell(200, 7, txt=f"{bullet}. Respostas possíveis:", ln=True)

            for idx_2, option in enumerate(options, start=1):
                if option.image:
                    y_start = pdf.get_y()
                    pdf.image(option.image.path, pdf.w / 2 - 30, y_start, 50, 50)
                    pdf.ln(60)
                else:
                    pdf.cell(25)
                    pdf.cell(200, 7, txt=f"{nested_bullet} {option.option}", ln=True)

            # if survey.answer:
            #     pdf.cell(15)
            #     pdf.cell(200, 7, txt=f"{bullet}. Sua resposta foi: {survey.answer.upper()}", ln=True)

            pdf.ln(14)
            pdf.cell(15)
            pdf.cell(200, 7, txt=f"{bullet}. RESULTADO:", ln=True)

            total_votes = options.aggregate(Sum('votes'))['votes__sum'] or 0
            pdf.cell(25)
            pdf.cell(200, 7, txt=f"{nested_bullet} Total de votos: {total_votes}", ln=True)

            for idx_2, option in enumerate(options, start=1):
                pdf.cell(25)
                if total_votes:
                    percent = ((option.votes * 100) / total_votes)
                else:
                    percent = 0
                pdf.cell(200, 7, txt=f"{option.option}: {option.votes} votos. Representando {round(percent, 2)} dos votos.", ln=True)

            if survey.location:

                pdf.cell(15)
                pdf.cell(200, 7, txt=f"Enquete solicitada por { survey.location.condominium.condominium_name }, em { survey.location.address }", ln=True)

            pdf.ln(7)
            line_start = pdf.get_y()
            pdf.line(25, line_start, 185, line_start)
            pdf.ln(15)

        messages.success(request, "Documento Gerado!")
        return _make_pdf_response(pdf, file_name)
    messages.success(request, "Nenhuma enquete realizada neste período!")
    return redirect(reverse('info:reports'))


@login_required(login_url='info:sign-in')
def export_contracts(request):
    condominium = get_condominium(request)
    form = GeneralReportForm()
    form.fields['initial'].label = "Que serão comunicadas de"

    if request.method == "POST":
        initial = request.POST.get('initial')
        until = request.POST.get('until')

        request.session['initial'] = initial
        request.session['until'] = until

        if parse_date(until) == date.today():
            until_filter = date.today() + timedelta(1)
        else:
            until_filter = until

        if initial == until:
            until_filter = parse_date(initial) + timedelta(1)

        objects_next_days = Contract.objects.filter(created__gte=initial, created__lte=until_filter,
                                                    condominium=condominium) \
            .order_by('notify_day')

        return _contracts_pdf(request, condominium, objects_next_days, initial, until)
    context = {'form': form,
               }
    return render(request, "info/condominium/report/contracts_report.html", context=context)


def _contracts_pdf(request, condominium, contracts, initial, until):
    if len(contracts):

        pdf = PDF(test=condominium.is_testing)
        pdf.add_page()
        pdf.set_auto_page_break(auto=True, margin=15)

        _add_pdf_header(pdf, condominium)

        pdf.set_font("helvetica", "B", 12)
        if initial != 0:
            file_name = "Relatório_Vencimentos"
            pdf.cell(0, 10,
                     txt=f"Relatório dos vencimentos de {datetime.strptime(initial, '%Y-%m-%d').strftime('%d/%m/%Y')} até {datetime.strptime(until, '%Y-%m-%d').strftime('%d/%m/%Y')}",
                     align="C", ln=True)
        else:
            file_name = "Vencimento_" + str(contracts[0].pk)

        pdf.ln(15)
        bullet = "-"

        for idx, contract in enumerate(contracts, start=1):
            pdf.cell(15)
            pdf.cell(200, 7, txt=f"{idx}. Nome: {contract.item}", ln=True)

            if contract.description:
                pdf.cell(15)
                pdf.cell(200, 7, txt=f"{bullet}. Descrição: {contract.description}", ln=True)

            if contract.last_maintenance:
                pdf.cell(15)
                pdf.cell(200, 7, txt=f"{bullet}. Última manutenção em: {contract.last_maintenance.strftime('%d/%m/%y')}", ln=True)
            if contract.next_maintenance:
                pdf.cell(15)
                pdf.cell(200, 7,
                         txt=f"{bullet}. Próxima manutenção em: {contract.next_maintenance.strftime('%d/%m/%y')}",
                         ln=True)
            pdf.cell(15)
            pdf.cell(200, 7, txt=f"{bullet}. Email que será notificado: {contract.to_email}", ln=True)
            pdf.cell(15)
            pdf.cell(200, 7, txt=f"{bullet}. Data da notificação: {contract.notify_day.strftime('%d/%m/%y')}", ln=True)

            if contract.image:
                y_start = pdf.get_y()
                pdf.image(contract.image.path, pdf.w / 2 - 30, y_start, 50, 50)
                pdf.ln(60)

            pdf.ln(7)
            line_start = pdf.get_y()
            pdf.line(25, line_start, 185, line_start)
            pdf.ln(15)

        messages.success(request, "Documento Gerado!")
        return _make_pdf_response(pdf, file_name)
    messages.success(request, "Nenhuma vencimento registrado neste período!")
    return redirect(reverse('info:reports'))


@login_required(login_url='info:sign-in')
def export_locations(request):
    condominium = get_condominium(request)
    form = GeneralReportForm()

    if request.method == "POST":
        initial = request.POST.get('initial')
        until = request.POST.get('until')

        request.session['initial'] = initial
        request.session['until'] = until

        if parse_date(until) == date.today():
            until_filter = date.today() + timedelta(1)
        else:
            until_filter = until

        if initial == until:
            until_filter = parse_date(initial) + timedelta(1)

        start_datetime = make_aware(datetime.combine(parse_date(initial), datetime.min.time()))
        end_datetime = make_aware(datetime.combine(parse_date(until_filter), datetime.min.time()))
        objects_last_days = UserLocation.objects.filter(created__gte=start_datetime, created__lte=end_datetime,
                                                        condominium=condominium).order_by('-created')

        return _locations_pdf(request, condominium, objects_last_days, initial, until)
    context = {'form': form,
               }
    return render(request, "info/condominium/report/locations_report.html", context=context)


def _locations_pdf(request, condominium, locations, initial, until):
    if len(locations) > 0:

        pdf = PDF(test=condominium.is_testing)
        pdf.add_page()
        pdf.set_auto_page_break(auto=True, margin=15)

        _add_pdf_header(pdf, condominium)

        pdf.set_font("helvetica", "B", 12)

        file_name = "Relatório_Localizações"
        pdf.cell(0, 10,
                 txt=f"Relatório das localizações de {datetime.strptime(initial, '%Y-%m-%d').strftime('%d/%m/%Y')} até {datetime.strptime(until, '%Y-%m-%d').strftime('%d/%m/%Y')}",
                 align="C", ln=True)

        pdf.ln(15)
        bullet = "-"

        for idx, location in enumerate(locations, start=1):
            pdf.cell(15)
            pdf.cell(200, 7, txt=f"{idx}. Localização:", ln=True)
            pdf.cell(25)
            pdf.cell(200, 7, txt=f"{bullet}. Usuário: {location.condominium.condominium_name}", ln=True)
            pdf.cell(25)
            pdf.cell(200, 7, txt=f"{bullet}. IP: {location.ip_address}", ln=True)
            pdf.cell(25)
            pdf.cell(200, 7, txt=f"{bullet}. Endereço: {location.address}", ln=True)
            pdf.cell(25)
            pdf.cell(200, 7, txt=f"{bullet}. Latitude: {location.latitude}", ln=True)
            pdf.cell(25)
            pdf.cell(200, 7, txt=f"{bullet}. Longitude: {location.longitude}", ln=True)
            pdf.cell(25)
            pdf.cell(200, 7, txt=f"{bullet}. Data: {location.created.astimezone(FIXED_TZ).strftime('%d/%m/%y - %H:%M:%S')}", ln=True)
            pdf.ln(7)
            line_start = pdf.get_y()
            pdf.line(25, line_start, 185, line_start)
            pdf.ln(15)

        messages.success(request, "Documento Gerado!")
        return _make_pdf_response(pdf, file_name)
    messages.success(request, "Nenhuma localização registrada neste período!")
    return redirect(reverse('info:reports'))


@login_required(login_url='info:sign-in')
def export_residents(request):
    condominium = get_condominium(request)
    residents = Resident.objects.filter(apartment__block__condominium=condominium)
    return _residents_pdf(request, condominium, residents)


def _residents_pdf(request, condominium, residents):
    pdf = PDF(test=condominium.is_testing)
    pdf.add_page()
    pdf.set_auto_page_break(auto=True, margin=15)

    _add_pdf_header(pdf, condominium)

    pdf.set_font("helvetica", "B", 12)

    file_name = "Relatório_Moradores"
    pdf.cell(0, 10,
             txt=f"Relatório dos moradores",
             align="C", ln=True)
    pdf.ln(15)
    bullet = "-"

    for idx, resident in enumerate(residents, start=1):
        pdf.cell(15)
        pdf.cell(200, 7, txt=f"{idx}. Morador:", ln=True)
        pdf.cell(25)
        pdf.cell(200, 7, txt=f"{bullet}. Bloco: {resident.apartment.block}", ln=True)
        pdf.cell(25)
        pdf.cell(200, 7, txt=f"{bullet}. Apartamento: {resident.apartment.number}", ln=True)
        pdf.cell(25)
        pdf.cell(200, 7, txt=f"{bullet}. Complement: {resident.apartment.complement}", ln=True)
        pdf.cell(25)
        pdf.cell(200, 7, txt=f"{bullet}. Nome: {resident.name}", ln=True)
        pdf.cell(25)
        pdf.cell(200, 7, txt=f"{bullet}. Tipo: {resident.kind}", ln=True)
        pdf.cell(25)
        pdf.cell(200, 7, txt=f"{bullet}. Email: {resident.email}", ln=True)
        pdf.ln(7)
        line_start = pdf.get_y()
        pdf.line(25, line_start, 185, line_start)
        pdf.ln(15)

    messages.success(request, "Documento Gerado!")
    return _make_pdf_response(pdf, file_name)


@login_required(login_url='info:sign-in')
def export_answer_to_pdf(request, id):
    condominium = get_condominium(request)
    answer = [ReviewAnswer.objects.get(pk=id)]
    return _answer_pdf(request, condominium, answer, 0)


def _answer_pdf(request, condominium, answer, days):
    pdf = PDF(test=condominium.is_testing)
    pdf.add_page()
    pdf.set_auto_page_break(auto=True, margin=15)

    _add_pdf_header(pdf, condominium)

    pdf.set_font("helvetica", "B", 12)

    if days == 0:
        file_name = "Avaliação_" + str(answer[0].pk)
        pdf.cell(0, 10,
                 txt=f"Avaliação do morador",
                 align="C", ln=True)
    else:
        file_name = "Relatório_Avaliações"
        pdf.cell(0, 10,
                 txt=f"Avaliações",
                 align="C", ln=True)

    pdf.ln(15)
    bullet = "-"

    for idx, ans in enumerate(answer, start=1):
        pdf.cell(15)
        pdf.cell(200, 7, txt=f"{idx}. Avaliação:", ln=True)
        pdf.cell(25)
        pdf.cell(200, 7, txt=f"{bullet}. Nome: {ans.name}", ln=True)
        if ans.email:
            pdf.cell(25)
            pdf.cell(200, 7, txt=f"{bullet}. Email: {ans.email}", ln=True)
        if ans.address:
            pdf.cell(25)
            pdf.cell(200, 7, txt=f"{bullet}. Endereço: {ans.address}", ln=True)
        if ans.message:
            pdf.cell(25)
            pdf.cell(200, 7, txt=f"{bullet}. Comentário: {ans.message}", ln=True)
        if ans.rate:
            pdf.cell(25)
            pdf.cell(200, 7, txt=f"{bullet}. Nota: {ans.rate}", ln=True)
        if ans.image:
            y_start = pdf.get_y()
            pdf.image(ans.image.path, pdf.w / 2 - 30, y_start, 50, 50)
            pdf.ln(60)
        pdf.cell(25)
        pdf.cell(200, 7, txt=f"{bullet}. Respondeu em: {ans.created.astimezone(FIXED_TZ).strftime('%d/%m/%y - %H:%M:%S')}", ln=True)
        pdf.ln(7)
        line_start = pdf.get_y()
        pdf.line(25, line_start, 185, line_start)
        pdf.ln(15)

    messages.success(request, "Documento Gerado!")
    return _make_pdf_response(pdf, file_name)


@login_required(login_url='info:sign-in')
def reviews_report(request):
    condominium = get_condominium(request)
    form = GeneralReportForm()

    if request.method == "POST":
        initial = request.POST.get('initial')
        until = request.POST.get('until')

        request.session['initial'] = initial
        request.session['until'] = until

        if parse_date(until) == date.today():
            until_filter = date.today() + timedelta(1)
        else:
            until_filter = until

        if initial == until:
            until_filter = parse_date(initial) + timedelta(1)

        objects_last_days = Review.objects.filter(created__gte=initial, created__lte=until_filter,
                                                  condominium=condominium).order_by('-created')

        return _reviews_pdf(request, condominium, objects_last_days, initial, until)
    context = {'form': form,
               }
    return render(request, "info/condominium/report/reviews_report.html", context=context)


def _reviews_pdf(request, condominium, reviews, initial, until):
    if len(reviews) > 0:

        pdf = PDF(test=condominium.is_testing)
        pdf.add_page()
        pdf.set_auto_page_break(auto=True, margin=15)

        _add_pdf_header(pdf, condominium)

        pdf.set_font("helvetica", "B", 12)

        if initial == 0:
            file_name = "Avaliação_" + str(reviews[0].pk)
            pdf.cell(0, 10,
                     txt=f"Avaliação do morador",
                     align="C", ln=True)
        else:
            file_name = "Relatório_Avaliações"
            pdf.cell(0, 10,
                     txt=f"Avaliações",
                     align="C", ln=True)

        pdf.ln(15)
        bullet = "-"
        nested_bullet = ">"
        for idx, review in enumerate(reviews, start=1):
            answers = ReviewAnswer.objects.filter(review=review, is_valid=True)
            pdf.cell(15)
            pdf.cell(200, 7, txt=f"{idx}. Serviço: {review.service}", ln=True)
            pdf.cell(25)
            pdf.cell(200, 7, txt=f"{bullet}. Nota: {review.rate}", ln=True)
            pdf.cell(25)
            pdf.cell(200, 7, txt=f"{bullet}. Prestador: {review.provider}", ln=True)
            pdf.cell(25)
            pdf.cell(200, 7, txt=f"{bullet}. Respostas:", ln=True)
            for idx_2, answer in enumerate(answers, start=1):
                pdf.cell(35)
                pdf.cell(200, 7, txt=f"{nested_bullet} Nome: { answer.name }", ln=True)
                pdf.cell(35)
                pdf.cell(200, 7, txt=f"{nested_bullet} Email: { answer.email }", ln=True)
                pdf.cell(35)
                pdf.cell(200, 7, txt=f"{nested_bullet} Endereço: {answer.address}", ln=True)
                pdf.cell(35)
                pdf.cell(200, 7, txt=f"{nested_bullet} Nota: {answer.rate}", ln=True)
                pdf.cell(35)
                pdf.cell(200, 7, txt=f"{nested_bullet} Comentário: {answer.message}", ln=True)
                if answer.image:
                    y_start = pdf.get_y()
                    pdf.image(answer.image.path, pdf.w / 2 - 30, y_start, 50, 50)
                    pdf.ln(60)
            pdf.ln(7)
            line_start = pdf.get_y()
            pdf.line(25, line_start, 185, line_start)
            pdf.ln(15)

        messages.success(request, "Documento Gerado!")
        return _make_pdf_response(pdf, file_name)
    messages.success(request, "Nenhuma avaliação realizada neste período!")
    return redirect(reverse('info:reports'))


def visitant_report(request):
    condominium = get_condominium(request)
    form = VisitantReportForm()

    if request.method == "POST":
        visits_from = request.POST.get('visits_from')
        visits_until = request.POST.get('visits_until')
        block = request.POST.get('block')
        start_datetime, end_datetime = _build_report_datetime_range(visits_from, visits_until)

        request.session['initial'] = visits_from
        request.session['until'] = visits_until

        if block:
            objects = VisitantReport.objects.filter(visit_in__gte=start_datetime, condominium=condominium,
                                                    visit_in__lt=end_datetime, block__contains=block)
        else:
            objects = VisitantReport.objects.filter(visit_in__gte=start_datetime, condominium=condominium,
                                                visit_in__lt=end_datetime)

        type = request.POST.get('type')

        if type == '1':

            return _visitants_pdf(request, condominium, objects, visits_from, visits_until)
        else:
            return _xls_visitants_report(request, condominium, objects, visits_from, visits_until)

    context = {'form': form,
               }

    return render(request, "info/condominium/report/visitant_report.html", context=context)


def _visitants_pdf(request, condominium, visitants, visits_from, visits_until):
    if len(visitants) > 0:
        context = {'condominium': condominium,
                   'visitants': visitants,
                   'from': visits_from,
                   'until': visits_until,
                   'test': condominium.is_testing}

        pdf = PDF(test=condominium.is_testing)
        pdf.add_page()
        pdf.set_auto_page_break(auto=True, margin=15)

        _add_pdf_header(pdf, condominium)

        pdf.set_font("helvetica", "B", 12)
        pdf.cell(0, 10, txt=f"Relatório dos Visitantes de {datetime.strptime(visits_from, '%Y-%m-%d').strftime('%d/%m/%Y')} até {datetime.strptime(visits_until, '%Y-%m-%d').strftime('%d/%m/%Y')}", align="C", ln=True)

        pdf.set_font("helvetica", "", 10)
        pdf.cell(145, 7, txt="Visitas:", ln=False)
        pdf.cell(50, 7, txt=f"Total de visitantes: {len(visitants)}", ln=True)

        bullet = "-"
        for idx, obj in enumerate(visitants, start=1):
            pdf.cell(15)
            pdf.cell(200, 7, txt=f"{idx}. Visitante: {obj.name}", ln=True)
            y_start = pdf.get_y()
            pdf.cell(25)
            pdf.cell(200, 7, txt=f"{bullet} Bloco: {obj.block}", ln=True)
            pdf.cell(25)
            pdf.cell(200, 7, txt=f"{bullet} Apartamento: {obj.apartment}", ln=True)
            if obj.resident:
                pdf.cell(25)
                pdf.cell(200, 7, txt=f"{bullet} Morador: {obj.resident.condominium_name}", ln=True)
            if obj.until:
                pdf.cell(25)
                pdf.cell(200, 7, txt=f"{bullet} Liberado até: {obj.until.strftime('%H:%M de %d/%m/%Y')}", ln=True)
            pdf.cell(25)
            pdf.cell(200, 7, txt=f"{bullet} Porteiro que liberou: {obj.security_name}", ln=True)
            if obj.visit_in:
                pdf.cell(25)
                pdf.cell(200, 7, txt=f"{bullet} Chegada: {obj.visit_in.astimezone(FIXED_TZ).strftime('%H:%M de %d/%m/%Y')}", ln=True)
            if obj.leaves_in:
                pdf.cell(25)
                pdf.cell(200, 7, txt=f"{bullet} Saída: {obj.leaves_in.astimezone(FIXED_TZ).strftime('%H:%M de %d/%m/%Y')}", ln=True)
            pdf.cell(25)
            pdf.cell(200, 7, txt=f"{bullet} Documento: {obj.document}", ln=True)
            pdf.cell(25)
            pdf.cell(200, 7, txt=f"{bullet} Modelo do veículo: {obj.vehicle_model}", ln=True)
            if obj.vehicle_plate:
                pdf.cell(25)
                pdf.cell(200, 7, txt=f"{bullet} Placa do veículo: {obj.vehicle_plate}", ln=True)
            if obj.photo:
                file_path = obj.photo.path
                if os.path.exists(file_path):
                    pdf.image(obj.photo.path, pdf.w - 115, y_start, 100)
            pdf.ln(7)
            line_start = pdf.get_y()
            pdf.line(25, line_start, 185, line_start)
            pdf.ln(15)

        return _make_pdf_response(pdf, "Relatório_Visitantes")

    messages.success(request, "Nenhuma visita realizada neste período!")
    return redirect(reverse('info:reports'))


def storage_report(request):
    condominium = get_condominium(request)
    form = GeneralReportForm()

    if request.method == "POST":
        initial = request.POST.get('initial')
        until = request.POST.get('until')

        request.session['initial'] = initial
        request.session['until'] = until

        if parse_date(until) == date.today():
            until_filter = date.today() + timedelta(1)
        else:
            until_filter = until

        if initial == until:
            until_filter = parse_date(initial) + timedelta(1)

        entries = []
        products = Product.objects.filter(condominium=condominium)
        for product in products:
            most_recent_adds = StorageEntry.objects.filter(created__gte=initial, product=product,
                                                           created__lte=until_filter, type="ENTRADA").order_by("-created")[:4]
            most_recent_withdraws = StorageEntry.objects.filter(created__gte=initial, product=product,
                                                                created__lte=until_filter, type="SAÍDA").order_by("-created")[
                                    :4]
            if len(most_recent_adds) or len(most_recent_withdraws):
                avg = 0
                date_initial = parse_date(initial)
                date_until = parse_date(until)

                time_delta = date_until - date_initial
                days = time_delta.days
                if len(most_recent_adds) > 1:
                    items = 0
                    for add in most_recent_adds:
                        items = items + add.quantity
                    item_avg = items / float(len(most_recent_adds))
                    avg = format(item_avg / days, ".2f")
                elif len(most_recent_adds) == 1:
                    add = most_recent_adds.first()
                    old_add = StorageEntry.objects.filter(created__lte=add.created,
                                                          product=product,
                                                          type="ENTRADA").order_by("-created").first()

                    time_delta = add.created.date() - old_add.created.date()
                    days = time_delta.days

                    avg = format(float(add.quantity) / days, ".2f")
                elif len(most_recent_withdraws) > 0:
                    items = 0
                    for wtdw in most_recent_withdraws:
                        items = items + wtdw.quantity
                    item_avg = items / float(len(most_recent_withdraws))
                    avg = format(item_avg / days, ".2f")
                else:
                    wtdw = most_recent_withdraws.first()
                    old_wtdw = StorageEntry.objects.filter(created__lte=wtdw.created,
                                                           product=product,
                                                           type="SAÍDA").order_by("-created").first()

                    time_delta = wtdw.created.date() - old_wtdw.created.date()
                    days = time_delta.days

                    avg = format(float(wtdw.quantity) / days, ".2f")

                entry = {
                    'product': product,
                    'avg': avg,
                }

                if len(most_recent_adds):
                    entry['recent_adds'] = most_recent_adds
                if len(most_recent_withdraws):
                    entry['recent_withdraws'] = most_recent_withdraws
                entries.append(entry)

        return _storage_pdf(request, condominium, entries, initial, until)
    context = {'form': form,
               }
    return render(request, "info/condominium/report/visitant_report.html", context=context)


def _storage_pdf(request, condominium, entries, visits_from, visits_until):
    if len(entries) > 0:

        pdf = PDF(test=condominium.is_testing)
        pdf.add_page()
        pdf.set_auto_page_break(auto=True, margin=15)

        _add_pdf_header(pdf, condominium)

        pdf.set_font("helvetica", "B", 12)


        file_name = "Relatório_Estoque"
        pdf.cell(0, 10,
                 txt=f"Relatório do estoque de {datetime.strptime(visits_from, '%Y-%m-%d').strftime('%d/%m/%Y')} até {datetime.strptime(visits_until, '%Y-%m-%d').strftime('%d/%m/%Y')}",
                 align="C", ln=True)

        pdf.ln(15)
        bullet = "-"
        nested_bullet = ">"
        for idx, entry in enumerate(entries, start=1):
            pdf.cell(15)
            pdf.cell(200, 7, txt=f"{idx}. Produto: {entry['product'].name}", ln=True)
            pdf.cell(25)
            pdf.cell(200, 7, txt=f"{bullet}. Descrição: {entry['product'].descrition}", ln=True)
            pdf.cell(25)
            pdf.cell(200, 7, txt=f"{bullet}. Quantidade em estoque: {entry['product'].quantity}", ln=True)
            pdf.cell(25)
            pdf.cell(200, 7, txt=f"{bullet}. Quantidade mínima: {entry['product'].warning_quantity}", ln=True)
            pdf.cell(25)
            pdf.cell(200, 7, txt=f"{bullet}. Duração média de unidades: { entry['avg'] } por dia", ln=True)
            if entry['product'].image:
                y_start = pdf.get_y()
                pdf.image(entry['product'].image.path, pdf.w / 2 - 30, y_start, 50, 50)
                pdf.ln(60)
            if entry['recent_adds']:
                pdf.cell(25)
                pdf.cell(200, 7, txt=f"{bullet}. Últimas entradas:", ln=True)
                for idx_2, add in enumerate(entry['recent_adds'], start=1):
                    pdf.cell(35)
                    pdf.cell(200, 7,
                             txt=f"{nested_bullet}: Dia e hora: { add.created.astimezone(FIXED_TZ).strftime('%d/%m/%y - %H:%M:%S') }",
                             ln=True)
                    pdf.cell(35)
                    pdf.cell(200, 7,
                             txt=f"{nested_bullet}: Quantidade: {add.quantity}",
                             ln=True)
                    pdf.cell(35)
                    pdf.cell(200, 7,
                             txt=f"{nested_bullet}: Preço: R$ {add.price}0",
                             ln=True)
                    pdf.cell(35)
                    pdf.cell(200, 7,
                             txt=f"{nested_bullet}: Funcionário que deu entrada: { add.worker.condominium_name }",
                             ln=True)
            if entry['recent_withdraws']:
                for idx_2, wit in enumerate(entry['recent_withdraws'], start=1):
                    pdf.cell(35)
                    pdf.cell(200, 7,
                             txt=f"{nested_bullet}: Dia e hora: {wit.created.astimezone(FIXED_TZ).strftime('%d/%m/%y - %H:%M:%S')}",
                             ln=True)
                    pdf.cell(35)
                    pdf.cell(200, 7,
                             txt=f"{nested_bullet}: Quantidade: {wit.quantity}",
                             ln=True)
                    pdf.cell(35)
                    pdf.cell(200, 7,
                             txt=f"{nested_bullet}: Funcionário que retirou: {wit.worker.condominium_name}",
                             ln=True)

            pdf.ln(7)
            line_start = pdf.get_y()
            pdf.line(25, line_start, 185, line_start)
            pdf.ln(15)

        messages.success(request, "Documento Gerado!")
        return _make_pdf_response(pdf, file_name)
    messages.success(request, "Nenhuma entrada ou retirada no estoque neste período!")
    return redirect(reverse('info:reports'))


def timeline_report(request):
    condominium = get_condominium(request)
    form = GeneralReportForm()

    if request.method == "POST":
        initial = request.POST.get('initial')
        until = request.POST.get('until')

        request.session['initial'] = initial
        request.session['until'] = until

        if parse_date(until) == date.today():
            until_filter = date.today() + timedelta(1)
        else:
            until_filter = until

        if initial == until:
            until_filter = parse_date(initial) + timedelta(1)

        objects = Timeline.objects.filter(created__gte=initial, condominium=condominium,
                                          created__lte=until_filter)

        return _timeline_pdf(request, condominium, objects, initial, until)
    context = {'form': form,
               }
    return render(request, "info/condominium/report/timeline_report.html", context=context)


def _timeline_pdf(request, condominium, objects, initial, until):
    if len(objects) > 0:
        pdf = PDF(test=condominium.is_testing)
        pdf.add_page()
        pdf.set_auto_page_break(auto=True, margin=15)

        _add_pdf_header(pdf, condominium)

        pdf.set_font("helvetica", "B", 12)

        if initial != 0:
            file_name = "Relatório_Timeline"
            pdf.cell(0, 10,
                     txt=f"Relatório de linhas do tempo de {datetime.strptime(initial, '%Y-%m-%d').strftime('%d/%m/%Y')} até {datetime.strptime(until, '%Y-%m-%d').strftime('%d/%m/%Y')}",
                     align="C", ln=True)
        else:
            file_name = "Timeline_" + str(objects[0].pk)

        pdf.ln(15)
        bullet = "-"
        nested_bullet = ">"
        for idx, timeline in enumerate(objects, start=1):
            phases = TimelinePhase.objects.filter(timeline=timeline).order_by('end_date')
            pdf.cell(15)
            pdf.cell(200, 7, txt=f"{idx}. Título: {timeline.title}", ln=True)
            pdf.cell(25)
            pdf.cell(200, 7, txt=f"{bullet}. Descrição: {timeline.description}", ln=True)
            pdf.cell(25)
            pdf.cell(200, 7, txt=f"{bullet}. Início: {timeline.start_date.strftime('%d/%m/%Y')}", ln=True)
            pdf.cell(25)
            pdf.cell(200, 7, txt=f"{bullet}. Fim: {timeline.end_date.strftime('%d/%m/%Y')}", ln=True)
            pdf.cell(25)
            pdf.cell(200, 7, txt=f"{bullet}. Etapas:", ln=True)

            for idx_2, phase in enumerate(phases, start=1):
                pdf.cell(35)
                pdf.cell(200, 7, txt=f"{nested_bullet} { phase.title }", ln=True)
                pdf.cell(35)
                pdf.cell(200, 7, txt=f"{nested_bullet} Descrição: {phase.description}", ln=True)
                pdf.cell(35)
                pdf.cell(200, 7, txt=f"{nested_bullet} Data de fim: {phase.end_date.strftime('%d/%m/%Y')}", ln=True)
                if phase.image:
                    y_start = pdf.get_y()
                    pdf.image(phase.image.path, pdf.w / 2 - 30, y_start, 50, 50)
                    pdf.ln(60)
                if phase.link:
                    pdf.cell(35)
                    pdf.set_text_color(0, 0, 255)
                    pdf.cell(200, 7, txt=f"{get_current_site(request).domain + phase.link}", ln=True,
                             link="http://" + get_current_site(request).domain + phase.link)
                    pdf.set_text_color(0, 0, 0)

            pdf.ln(7)
            line_start = pdf.get_y()
            pdf.line(25, line_start, 185, line_start)
            pdf.ln(15)

        messages.success(request, "Documento Gerado!")
        return _make_pdf_response(pdf, file_name)
    messages.success(request, "Nenhuma timeline criada neste período!")
    return redirect(reverse('info:reports'))


def resident_activity_report(request):
    condominium = get_condominium(request)
    form = GeneralReportForm()

    if request.method == "POST":
        initial = request.POST.get('initial')
        until = request.POST.get('until')

        request.session['initial'] = initial
        request.session['until'] = until

        if parse_date(until) == date.today():
            until_filter = date.today() + timedelta(1)
        else:
            until_filter = until

        if initial == until:
            until_filter = parse_date(initial) + timedelta(1)

        objects = ResidentActivity.objects.filter(created__gte=initial, condominium=condominium,
                                                  created__lte=until_filter)

        return _resident_activity_pdf(request, condominium, objects, initial, until)
    context = {'form': form,
               }
    return render(request, "info/condominium/report/user_activity_report.html", context=context)


def _resident_activity_pdf(request, condominium, objects, initial, until):
    if len(objects) > 0:

        pdf = PDF(test=condominium.is_testing)
        pdf.add_page()
        pdf.set_auto_page_break(auto=True, margin=15)

        _add_pdf_header(pdf, condominium)

        pdf.set_font("helvetica", "B", 12)

        if initial != 0:
            file_name = "Relatório_Solicitações"
            pdf.cell(0, 10,
                     txt=f"Relatório das solicitações de {datetime.strptime(initial, '%Y-%m-%d').strftime('%d/%m/%Y')} até {datetime.strptime(until, '%Y-%m-%d').strftime('%d/%m/%Y')}",
                     align="C", ln=True)
        else:
            file_name = "Solicitação_" + str(objects[0].pk)

        pdf.ln(15)
        bullet = "-"
        nested_bullet = ">"
        for idx, activity in enumerate(objects, start=1):
            answers = ResidentActivityAnswer.objects.filter(activity=activity).order_by('created')
            pdf.cell(15)
            pdf.cell(200, 7, txt=f"{idx}. Título: {activity.title}", ln=True)
            pdf.cell(25)
            pdf.cell(200, 7, txt=f"{bullet}. Tipo: {activity.kind}", ln=True)
            pdf.cell(25)
            pdf.cell(200, 7, txt=f"{bullet}. Descrição: {activity.description}", ln=True)
            if activity.resident:
                pdf.cell(25)
                pdf.cell(200, 7, txt=f"{bullet}. Morador: {activity.resident}", ln=True)
            if activity.apartment:
                pdf.cell(25)
                pdf.cell(200, 7, txt=f"{bullet}. Apartamento: {activity.apartment.number} {activity.apartment.complement}", ln=True)
            if activity.responsible:
                pdf.cell(25)
                pdf.cell(200, 7, txt=f"{bullet}. Responsável: {activity.responsible}", ln=True)
            pdf.cell(25)
            pdf.cell(200, 7, txt=f"{bullet}. Protocolo: {activity.protocol}", ln=True)
            pdf.cell(25)
            pdf.cell(200, 7, txt=f"{bullet}. Situação: {activity.status.upper()}", ln=True)
            pdf.cell(25)
            pdf.cell(200, 7, txt=f"{nested_bullet} Criada em: {activity.created.astimezone(FIXED_TZ).strftime('%d/%m/%Y')}", ln=True)
            pdf.cell(25)
            pdf.cell(200, 7, txt=f"{nested_bullet} Pespostas:", ln=True)
            for idx_2, answer in enumerate(answers, start=1):
                pdf.cell(35)
                pdf.cell(200, 7, txt=f"Autor: {answer.auteur}", ln=True)
                pdf.cell(35)
                pdf.cell(200, 7, txt=f"Mensagem: {answer.message}", ln=True)

            pdf.ln(7)
            line_start = pdf.get_y()
            pdf.line(25, line_start, 185, line_start)
            pdf.ln(15)

        messages.success(request, "Documento Gerado!")
        return _make_pdf_response(pdf, file_name)
    messages.success(request, "Nenhuma solicitação feita neste período!")
    return redirect(reverse('info:reports'))


def survey_votes_report(request, survey_id):

    condominium = get_condominium(request)
    survey = SurveyModel.objects.get(pk=survey_id)
    votes = SurveyAnswerModel.objects.filter(survey=survey).order_by('name')

    pdf = PDF(test=condominium.is_testing)
    pdf.add_page()
    pdf.set_auto_page_break(auto=True, margin=15)

    _add_pdf_header(pdf, condominium)

    pdf.set_font("helvetica", "B", 12)


    file_name = "Relatório_Votos"
    pdf.cell(0, 10,
             txt=f"VOTOS",
             align="C", ln=True)

    pdf.ln(15)
    bullet = "-"
    nested_bullet = ">"

    for idx, vote in enumerate(votes, start=1):
        pdf.cell(15)
        pdf.cell(200, 7, txt=f"{idx} Voto:", ln=True)
        pdf.cell(25)
        pdf.cell(200, 7, txt=f"{bullet} Nome: {vote.name}", ln=True)
        pdf.cell(25)
        pdf.cell(200, 7, txt=f"{bullet} Endereço: {vote.address}", ln=True)
        pdf.cell(25)
        pdf.cell(200, 7, txt=f"{bullet} Voto: {vote.option.option}", ln=True)
        if vote.answer_pic:
            y_start = pdf.get_y()
            pdf.image(vote.answer_pic.path, pdf.w / 2 - 30, y_start, 50, 50)
            pdf.ln(60)

    pdf.cell(15)
    pdf.cell(200, 7, txt=f"Total de votos: {len(votes)}", ln=True)

    pdf.ln(7)
    line_start = pdf.get_y()
    pdf.line(25, line_start, 185, line_start)
    pdf.ln(15)

    messages.success(request, "Documento Gerado!")
    return _make_pdf_response(pdf, file_name)


def reservation_report(request):

    condominium = get_condominium(request)
    form = GeneralReportForm()

    if request.method == "POST":
        initial = request.POST.get('initial')
        until = request.POST.get('until')

        request.session['initial'] = initial
        request.session['until'] = until

        if parse_date(until) == date.today():
            until_filter = date.today() + timedelta(1)
        else:
            until_filter = until

        if initial == until:
            until_filter = parse_date(initial) + timedelta(1)

        objects = Reservation.objects.filter(created__gte=initial, condominium=condominium,
                                          created__lte=until_filter)

        return export_reservation(request, condominium, objects, initial, until)
    context = {'form': form,
               }
    return render(request, "info/condominium/report/reservation_report.html", context=context)


def _reservation_pdf(request, condominium, objects, initial, until):

    if len(objects) > 0:

        pdf = PDF(test=condominium.is_testing)
        pdf.add_page()
        pdf.set_auto_page_break(auto=True, margin=15)

        _add_pdf_header(pdf, condominium)

        pdf.set_font("helvetica", "B", 12)

        if initial != 0:
            file_name = "Relatório_Reservas"
            pdf.cell(0, 10,
                     txt=f"Relatório das reservas de {datetime.strptime(initial, '%Y-%m-%d').strftime('%d/%m/%Y')} até {datetime.strptime(until, '%Y-%m-%d').strftime('%d/%m/%Y')}",
                     align="C", ln=True)
        else:
            file_name = "Reserva_" + str(objects[0].pk)

        pdf.ln(15)
        bullet = "-"
        nested_bullet = ">"
        for idx, booking in enumerate(objects, start=1):
            pdf.cell(15)
            pdf.cell(200, 7, txt=f"{idx}. Reserva:", ln=True)
            pdf.cell(25)
            pdf.cell(200, 7, txt=f"{bullet}. Área: {booking.place.name}", ln=True)
            pdf.cell(25)
            pdf.cell(200, 7, txt=f"{bullet}. Data: {booking.date.strftime('%d/%m/%Y')}", ln=True)
            pdf.cell(25)
            pdf.cell(200, 7, txt=f"{bullet}. Horário: {booking.time.init_time.strftime('%H:%M')} até {booking.time.end_time.strftime('%H:%M')}", ln=True)
            pdf.cell(25)
            pdf.cell(200, 7, txt=f"{bullet}. Situação: {booking.status}", ln=True)
            if booking.approved_by:
                pdf.cell(25)
                pdf.cell(200, 7, txt=f"{bullet}. Aprovado por: {booking.approved_by}", ln=True)
            if booking.canceled_by:
                pdf.cell(25)
                pdf.cell(200, 7, txt=f"{bullet}. Cancelado por: {booking.canceled_by}", ln=True)
            pdf.cell(25)
            pdf.cell(200, 7, txt=f"{bullet}. Criada em: {booking.created.astimezone(FIXED_TZ).strftime('%d/%m/%Y')}", ln=True)

        pdf.ln(7)
        line_start = pdf.get_y()
        pdf.line(25, line_start, 185, line_start)
        pdf.ln(15)

        messages.success(request, "Documento Gerado!")
        return _make_pdf_response(pdf, file_name)
    messages.success(request, "Nenhuma reserva realizada neste período!")
    return redirect(reverse('info:reports'))


def export_reservation(request, condominium, reservation_list, initial=None, until=None):
    if len(reservation_list) > 0:
        pdf = PDF(test=condominium.is_testing)
        pdf.add_page()
        pdf.set_auto_page_break(auto=True, margin=15)

        _add_pdf_header(pdf, condominium)

        pdf.set_font("helvetica", "B", 12)

        if initial and until:
            title = f"Relatório das Reservas de {datetime.strptime(initial, '%Y-%m-%d').strftime('%d/%m/%Y')} até {datetime.strptime(until, '%Y-%m-%d').strftime('%d/%m/%Y')}"
            filename = "Relatorio_Reservas"
        else:
            title = "Reserva"
            filename = "reserva"

        pdf.cell(0, 10,
                 title,
                 align="C", ln=True)

        pdf.set_font("helvetica", "", 10)
        bullet = "-"

        if initial and until:
            pdf.cell(145, 7, txt="Reservas:", ln=False)
            pdf.cell(50, 7, txt=f"Total de reservas: {len(reservation_list)}", ln=True)

        for idx, obj in enumerate(reservation_list, start=1):

            pdf.cell(25)
            pdf.cell(200, 7, txt=f"{bullet} Local: {obj.place.name}", ln=True)
            pdf.cell(25)
            pdf.cell(200, 7, txt=f"{bullet} Datada reserva: {obj.date.strftime('%d/%m/%Y')}", ln=True)
            pdf.cell(25)
            pdf.cell(200, 7, txt=f"{bullet} Início: {obj.time.init_time.strftime('%H:%M')}", ln=True)
            pdf.cell(25)
            pdf.cell(200, 7, txt=f"{bullet} Até: {obj.time.end_time.strftime('%H:%M')}", ln=True)
            pdf.cell(25)
            pdf.cell(200, 7, txt=f"{bullet} Morador: {obj.resident.condominium_name}", ln=True)
            pdf.cell(25)
            pdf.cell(200, 7, txt=f"{bullet} Situação: {obj.status}", ln=True)

            if obj.guests:
                pdf.cell(25)
                pdf.cell(200, 7, txt=f"{bullet} Convidados:", ln=True)
                pdf.cell(35)
                total = len(obj.guests)
                if total > 70:
                    missing = total
                    init = 0
                    while(missing > 0):
                        guests_aux = condominium.address[init : init + 70]

                        space_pos = guests_aux.rfind(' ', init, init + 69)
                        pdf.cell(200, 7,
                                       txt=f"{condominium.guests[init:space_pos]}",
                                       ln=True)
                        init = space_pos
                        missing = missing - len(condominium.guests[init:space_pos])
                else:
                    pdf.cell(200, 7, txt=f"{condominium.address}", ln=True)

            if obj.link:
                pdf.cell(25)
                pdf.set_text_color(0, 0, 255)
                pdf.cell(200, 7, txt=f"{get_current_site(request).domain + obj.link}", ln=True,
                         link="http://" + get_current_site(request).domain + obj.link)

            if obj.bill:
                pdf.cell(35)
                pdf.set_text_color(0, 0, 255)  # Set text color to blue (optional)
                pdf.set_font(style='U')  # Underline the text (optional)
                pdf.cell(0, 10, f"{get_current_site(request).domain + obj.bill.url}", ln=True,
                         link="http://" + get_current_site(request).domain + obj.bill.url)

            if obj.payment:
                pdf.cell(35)
                pdf.set_text_color(0, 0, 255)  # Set text color to blue (optional)
                pdf.set_font(style='U')  # Underline the text (optional)
                pdf.cell(0, 10, f"Comprovante: {get_current_site(request).domain + obj.payment.url}", ln=True,
                         link="http://" + get_current_site(request).domain + obj.payment.url)

            pdf.set_text_color(0, 0, 0)

            if obj.approved_by:
                pdf.cell(25)
                pdf.cell(200, 7, txt=f"{bullet} Aprovado por: {obj.approved_by}", ln=True)

            if obj.canceled_by:
                pdf.cell(25)
                pdf.cell(200, 7, txt=f"{bullet} Cancelado por: {obj.canceled_by}", ln=True)

            pdf.cell(25)
            pdf.cell(200, 7, txt=f"{bullet} Criado em: {obj.created.astimezone(FIXED_TZ).strftime('%d/%m/%Y')}", ln=True)

            pdf.ln(7)
            line_start = pdf.get_y()
            pdf.line(25, line_start, 185, line_start)
            pdf.ln(15)

        return _make_pdf_response(pdf, filename)
    messages.success(request, "Nenhuma reserca neste período!")
    return redirect(reverse('info:dashboard'))



@login_required(login_url='info:sign-in')
def export_user_control(request):
    condominium = get_condominium(request)
    form = GeneralReportForm()

    if request.method == "POST":
        initial = request.POST.get('initial')
        until = request.POST.get('until')

        request.session['initial'] = initial
        request.session['until'] = until

        if parse_date(until) == date.today():
            until_filter = date.today() + timedelta(1)
        else:
            until_filter = until

        if initial == until:
            until_filter = parse_date(initial) + timedelta(1)

        start_datetime = make_aware(datetime.combine(parse_date(initial), datetime.min.time()))
        end_datetime = make_aware(datetime.combine(until_filter, datetime.min.time()))
        objects_last_days = UserControl.objects.filter(created__gte=start_datetime, created__lte=end_datetime,
                                                        condominium=condominium).order_by('-created')

        return _control_pdf(request, condominium, objects_last_days, initial, until)
    context = {'form': form,
               }
    return render(request, "info/condominium/report/control_report.html", context=context)


def _control_pdf(request, condominium, locations, initial, until):
    if len(locations) > 0:

        pdf = PDF(test=condominium.is_testing)
        pdf.add_page()
        pdf.set_auto_page_break(auto=True, margin=15)

        _add_pdf_header(pdf, condominium)

        pdf.set_font("helvetica", "B", 12)

        if initial != 0:
            file_name = "Relatório_Sessões"
            pdf.cell(0, 10,
                     txt=f"Relatório das sessões de {datetime.strptime(initial, '%Y-%m-%d').strftime('%d/%m/%Y')} até {datetime.strptime(until, '%Y-%m-%d').strftime('%d/%m/%Y')}",
                     align="C", ln=True)
        else:
            file_name = "Reserva_" + str(locations[0].pk)

        pdf.ln(15)
        bullet = "-"
        nested_bullet = ">"
        for idx, control in enumerate(locations, start=1):
            pdf.cell(15)
            pdf.cell(200, 7, txt=f"{idx}. Usuário: {control.user.condominium_name}", ln=True)
            pdf.cell(25)
            pdf.cell(200, 7, txt=f"{bullet} IP: {control.ip_address}", ln=True)
            pdf.cell(25)
            pdf.cell(200, 7, txt=f"{bullet} Endereço: {control.address}", ln=True)
            pdf.cell(25)
            pdf.cell(200, 7, txt=f"{bullet} Criada em: {control.created.astimezone(FIXED_TZ).strftime('%d/%m/%Y')}", ln=True)
            pdf.cell(25)
            pdf.cell(200, 7, txt=f"{bullet} Check-in: {control.check_in.astimezone(FIXED_TZ).strftime('%d/%m/%Y - %H:%M:%S')}", ln=True)
            if control.check_out:
                pdf.cell(25)
                pdf.cell(200, 7, txt=f"{bullet} Check-out: {control.check_out.astimezone(FIXED_TZ).strftime('%d/%m/%Y - %H:%M:%S')}", ln=True)
                pdf.cell(25)
                pdf.cell(200, 7, txt=f"{bullet} Duração: {control.session_time}", ln=True)
            else:
                pdf.cell(25)
                pdf.cell(200, 7, txt=f"{bullet} Sessão ativa",
                         ln=True)

        pdf.ln(7)
        line_start = pdf.get_y()
        pdf.line(25, line_start, 185, line_start)
        pdf.ln(15)

        messages.success(request, "Documento Gerado!")
        return _make_pdf_response(pdf, file_name)
    messages.success(request, "Nenhuma sessão registrada neste período!")
    return redirect(reverse('info:reports'))


@login_required(login_url='info:sign-in')
def export_notification_to_pdf(request, id):
    condominium = get_condominium(request)

    message = [Message.objects.get(pk=id)]

    return _notification_pdf(request, condominium, message, message[0].kind, 0, 0, None)


def _notification_pdf(request, condominium, messages_list, kind, initial, until, resident):
    if len(messages_list) > 0:

        pdf = PDF(test=condominium.is_testing)
        pdf.add_page()
        pdf.set_auto_page_break(auto=True, margin=15)

        _add_pdf_header(pdf, condominium)

        pdf.set_font("helvetica", "B", 12)

        if initial and until:
            title = f"Relatório das Notificações de {datetime.strptime(initial, '%Y-%m-%d').strftime('%d/%m/%Y')} até {datetime.strptime(until, '%Y-%m-%d').strftime('%d/%m/%Y')}"
            filename = "Relatorio_Notificações"
        else:
            title = "Notificação"
            filename = "notificacao"

        pdf.cell(0, 10,
                 title,
                 align="C", ln=True)

        pdf.set_font("helvetica", "", 10)
        bullet = "-"

        if initial and until:
            pdf.cell(145, 7, txt="Notificações:", ln=False)
            pdf.cell(50, 7, txt=f"Total de notificações: {len(messages_list)}", ln=True)

        for idx, obj in enumerate(messages_list, start=1):
            files = MessageFileModel.objects.filter(message=obj)

            pdf.cell(25)
            pdf.cell(200, 7, txt=f"{bullet} Tipo: {obj.kind}", ln=True)

            if obj.block:
                pdf.cell(25)
                pdf.cell(200, 7, txt=f"{bullet} Para o Bloco: {obj.block}", ln=True)
            if obj.apartment:
                pdf.cell(25)
                pdf.cell(200, 7, txt=f"{bullet} Para o Apartamento: {obj.apartment}", ln=True)
            if obj.resident:
                pdf.cell(25)
                res = "Todos" if obj.resident == 'all' else obj.resident
                pdf.cell(200, 7, txt=f"{bullet} Para o Morador: {res}", ln=True)
            if obj.message:
                pdf.cell(25)
                pdf.cell(200, 7, txt=f"{bullet} Mensagem: {obj.message}", ln=True)

            pdf.cell(25)
            pdf.cell(200, 7, txt=f"{bullet} Data: {obj.created.astimezone(FIXED_TZ).strftime('%H:%M de %d/%m/%Y')}", ln=True)

            if files:
                pdf.cell(25)
                pdf.cell(200, 7, txt=f"{bullet} Anexos:", ln=True)
            for file in files:
                pdf.cell(35)
                pdf.set_text_color(0, 0, 255)  # Set text color to blue (optional)
                pdf.set_font(style='U')  # Underline the text (optional)
                pdf.cell(0, 10, f"{get_current_site(request).domain + file.file.url}", ln=True,
                         link="http://" + get_current_site(request).domain + file.file.url)

            pdf.set_text_color(0, 0, 0)
            pdf.ln(7)
            line_start = pdf.get_y()
            pdf.line(25, line_start, 185, line_start)
            pdf.ln(15)

        return _make_pdf_response(pdf, filename)
    messages.success(request, "Nenhuma notificação enviada neste período!")
    return redirect(reverse('info:reports'))



def export_to_exel(request, model):

    condominium = get_condominium(request)

    if 'initial' in request.session:
        initial = request.session['initial']
    else:
        initial = None

    if 'until' in request.session:
        until = request.session['until']
    else:
        until = None

    if 'kind' in request.session:
        kind = request.session['kind']
    else:
        kind = None

    if model == 'residents':
        queryset = Resident.objects.filter(apartment__block__condominium=condominium).order_by('apartment__block__name', 'apartment__number', 'apartment__complement')
        return _xls_resident_report(condominium, queryset)

    elif model == 'activities':

        if parse_date(until) == date.today():
            until_filter = date.today() + timedelta(1)
        else:
            until_filter = until

        if initial == until:
            until_filter = parse_date(initial) + timedelta(1)

        if not kind or kind == "TODAS":

            activities = Informative.objects.filter(created__gte=initial, created__lte=until_filter,
                                                           condominium=condominium)
            kind = "TODAS as atividades"
        else:
            informative_kind = InformativeKind.objects.get(pk=int(kind))
            activities = Informative.objects.filter(created__gte=initial, created__lte=until_filter,
                                                           condominium=condominium,
                                                           kind__iexact=informative_kind.name)

        return _xls_activity_report(request, condominium, activities, kind, initial, until)

    elif model == 'orders':

        orders = Order.objects.filter(created__gte=initial, created__lte=until,
                             apartment__block__condominium=condominium)

        return _xls_order_report(request, condominium, initial, until, orders)

    elif model == 'messages':

        if parse_date(until) == date.today():
            until_filter = date.today() + timedelta(1)
        else:
            until_filter = until

        if initial == until:
            until_filter = parse_date(initial) + timedelta(1)

        if kind == "TODOS":
            messages_to = Message.objects.filter(created__gte=initial, created__lte=until_filter,
                                                        condominium=condominium)
            kind = ""
        else:
            messages_to = Message.objects.filter(created__gte=initial, created__lte=until_filter,
                                                        condominium=condominium,
                                                        kind__iexact=kind)

        return _xls_messages_report(request, condominium, messages_to, kind, initial, until)

    elif model == 'reviews':

        if parse_date(until) == date.today():
            until_filter = date.today() + timedelta(1)
        else:
            until_filter = until

        if initial == until:
            until_filter = parse_date(initial) + timedelta(1)

        reviews = Review.objects.filter(created__gte=initial, created__lte=until_filter,
                                                  condominium=condominium).order_by('-created')

        return _xls_reviews_report(request, condominium, reviews, initial, until)

    elif model == 'surveys':

        if parse_date(until) == date.today():
            until_filter = date.today() + timedelta(1)
        else:
            until_filter = until

        if initial == until:
            until_filter = parse_date(initial) + timedelta(1)

        surveys = SurveyModel.objects.filter(created__gte=initial, created__lte=until_filter,
                                                       condominium=condominium)

        return _xls_surveys_report(request, condominium, surveys, initial, until)

    elif model == 'contracts':

        if parse_date(until) == date.today():
            until_filter = date.today() + timedelta(1)
        else:
            until_filter = until

        if initial == until:
            until_filter = parse_date(initial) + timedelta(1)

        contracts = Contract.objects.filter(notify_day__gte=initial, notify_day__lte=until_filter,
                                          condominium=condominium).order_by('notify_day')

        return _xls_contracts_report(request, condominium, contracts, initial, until)

    elif model == 'checklists':

        if parse_date(until) == date.today():
            until_filter = date.today() + timedelta(1)
        else:
            until_filter = until

        if initial == until:
            until_filter = parse_date(initial) + timedelta(1)

        checklists = Checklist.objects.filter(created__gte=initial, created__lte=until_filter,
                                                        condominium=condominium)

        return _xls_checklists_report(request, condominium, checklists, initial, until)

    elif model == 'locations':

        if parse_date(until) == date.today():
            until_filter = date.today() + timedelta(1)
        else:
            until_filter = until

        if initial == until:
            until_filter = parse_date(initial) + timedelta(1)

        start_datetime = make_aware(datetime.combine(parse_date(initial), datetime.min.time()))
        end_datetime = make_aware(datetime.combine(until_filter, datetime.min.time()))
        locations = UserLocation.objects.filter(created__gte=start_datetime, created__lte=end_datetime,
                                                        condominium=condominium).order_by('-created')

        return _xls_locations_report(request, condominium, locations, initial, until)

    elif model == 'visitants':

        if parse_date(until) == date.today():
            until_filter = date.today() + timedelta(1)
        else:
            until_filter = until

        if initial == until:
            until_filter = parse_date(initial) + timedelta(1)

        visitants = VisitantReport.objects.filter(visit_in__gte=initial, condominium=condominium,
                                                visit_in__lte=until_filter)

        return _xls_visitants_report(request, condominium, visitants, initial, until)

    elif model == 'timelines':

        if parse_date(until) == date.today():
            until_filter = date.today() + timedelta(1)
        else:
            until_filter = until

        if initial == until:
            until_filter = parse_date(initial) + timedelta(1)

        timelines = Timeline.objects.filter(created__gte=initial, condominium=condominium,
                                          created__lte=until_filter)

        return _xls_timelines_report(request, condominium, timelines, initial, until)

    elif model == 'user_activities':

        if parse_date(until) == date.today():
            until_filter = date.today() + timedelta(1)
        else:
            until_filter = until

        if initial == until:
            until_filter = parse_date(initial) + timedelta(1)

        resident_activities = ResidentActivity.objects.filter(created__gte=initial, condominium=condominium,
                                                  created__lte=until_filter)

        return _xls_resident_activities_report(request, condominium, resident_activities, initial, until)

    elif model == 'booking':

        if parse_date(until) == date.today():
            until_filter = date.today() + timedelta(1)
        else:
            until_filter = until

        if initial == until:
            until_filter = parse_date(initial) + timedelta(1)

        reservations = Reservation.objects.filter(created__gte=initial, condominium=condominium,
                                             created__lte=until_filter)

        return _xls_booking_report(request, condominium, reservations, initial, until)

    elif model == 'sessions':

        if parse_date(until) == date.today():
            until_filter = date.today() + timedelta(1)
        else:
            until_filter = until

        if initial == until:
            until_filter = parse_date(initial) + timedelta(1)

        start_datetime = make_aware(datetime.combine(parse_date(initial), datetime.min.time()))
        end_datetime = make_aware(datetime.combine(until_filter, datetime.min.time()))
        sessions = UserControl.objects.filter(created__gte=start_datetime, created__lte=end_datetime,
                                                       condominium=condominium).order_by('-created')

        return _xls_sessions_report(request, condominium, sessions, initial, until)

    elif model == 'storage':

        products = Product.objects.filter(condominium=condominium)

        return _xls_storage_report(request, condominium, products, initial, until)

    else:
        messages.error(request, "Relatório não suportado!")
        return redirect(reverse('info:reports'))


def _xls_resident_report(condominium, residents):
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml')
    response['Content-Disposition'] = 'attachment; filename="Relatorio_Moradores.xlsx"'

    workbook = Workbook()
    worksheet = workbook.active

    worksheet.merge_cells('A1:F1')
    worksheet.merge_cells('A2:F2')

    _add_xls_header(condominium, worksheet, "Relatório de Moradores")

    columns = ['Bloco', 'Apartamento', 'Complemento', 'Nome', 'Tipo', 'Email']

    row_num = 3

    _add_xls_columns(columns, row_num, worksheet)

    for resident in residents:
        row_num += 1

        row = [resident.apartment.block.name, resident.apartment.number, resident.apartment.complement, resident.name, resident.kind,
               resident.email]

        _add_cell_values(row, row_num, worksheet)

    workbook.save(response)
    return response


def _xls_activity_report(request, condominium, activities, kind, initial, until):
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml')
    response['Content-Disposition'] = 'attachment; filename="Relatorio_Atividades.xlsx"'

    workbook = Workbook()
    worksheet = workbook.active

    worksheet.merge_cells('A1:H1')
    worksheet.merge_cells('A2:H2')

    _add_xls_header(condominium, worksheet, "Relatório de Atividades", initial, until)

    columns = ['Titulo', 'Tipo', 'Funções', 'Descrição', 'Imagem', 'Arquivos', 'Video', 'Localização']

    row_num = 3

    current_site = get_current_site(request)

    _add_xls_columns(columns, row_num, worksheet)

    for activity in activities:
        row_num += 1
        title = activity.title
        atv_kind = activity.kind
        location = activity.location.address if activity.location else ""
        function_list = []
        function_desc = []
        function_images = []
        function_files = []
        function_videos = []

        functions = ActivityFunction.objects.filter(informative=activity)

        for func in functions:
            function_list.append(func.title)
            function_desc.append(func.description if func.description else "")
            function_videos = func.link
            images = ImageModel.objects.filter(function_item=func)

            if len(images) > 0:
                for image in images:
                    function_images.append("https://" + current_site.domain + image.image.url)

            files = FunctionItemFileModel.objects.filter(function_item=func)

            if len(files) > 0:
                for file in files:
                    function_files.append("https://" + current_site.domain + file.file.url)

        funcs = ','.join(str(atv_func) for atv_func in function_list)
        func_descs = ','.join(str(desc) for desc in function_desc)
        func_images = ','.join(str(im) for im in function_images)
        func_files = ','.join(str(fl) for fl in function_files)

        row = [title, atv_kind, funcs, func_descs,
               func_images, func_files, function_videos, location]

        _add_cell_values(row, row_num, worksheet)

    workbook.save(response)
    return response


def _xls_order_report(request, condominium, initial, until, orders):
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml')
    response['Content-Disposition'] = 'attachment; filename="Relatorio_Correspondencias.xlsx"'

    workbook = Workbook()
    worksheet = workbook.active

    worksheet.merge_cells('A1:J1')
    worksheet.merge_cells('A2:J2')

    _add_xls_header(condominium, worksheet, "Relatório de Correspondências", initial, until)

    columns = ['Nome', 'Descrição', 'Recebida por', 'Apartamento', 'Imagem', 'Código', 'Notificação', 'Entregue',
               'Entregue por', 'Retirada por']

    row_num = 3

    current_site = get_current_site(request)

    _add_xls_columns(columns, row_num, worksheet)

    for order in orders:
        row_num += 1
        descripition = order.description if order.description else ""
        apartment = str(order.apartment.number) + " " + order.apartment.complement if order.apartment else ""

        image = "https://" + current_site.domain + order.image.url if order.image else ""

        if order.delivered:
            delivery_date = order.delivered.astimezone(FIXED_TZ).strftime('%d/%m/%Y %H:%M')
            delivery_by = order.delivered_by
            collected_by = order.collected_by
        else:
            delivery_date = ""
            delivery_by = ""
            collected_by = ""

        row = [order.name, descripition, order.received_by, apartment, image, order.code,
               order.created.astimezone(FIXED_TZ).strftime('%d/%m/%Y %H:%M'), delivery_date, delivery_by, collected_by]

        _add_cell_values(row, row_num, worksheet)

    workbook.save(response)
    return response


def _xls_messages_report(request, condominium, messages, kind, initial, until):
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml')
    response['Content-Disposition'] = 'attachment; filename="Relatorio_Cominicados.xlsx"'

    workbook = Workbook()
    worksheet = workbook.active

    worksheet.merge_cells('A1:F1')
    worksheet.merge_cells('A2:F2')

    _add_xls_header(condominium, worksheet, "Relatório de Comunicados", initial, until)

    columns = ['Tipo', 'Bloco', 'Apartamento', 'Mensagem', 'Enviada em', 'Anexos']

    row_num = 3

    current_site = get_current_site(request)

    _add_xls_columns(columns, row_num, worksheet)

    for message in messages:

        row_num += 1
        file_list = []
        files = MessageFileModel.objects.filter(message=message)
        if len(files) > 0:
            for file in files:
                file_list.append("https://" + current_site.domain + file.file.url)

        message_files = ','.join(str(fl) for fl in file_list)

        row = [message.kind, message.block, message.apartment, message.message,
               message.created.astimezone(FIXED_TZ).strftime('%d/%m/%Y %H:%M'), message_files]

        _add_cell_values(row, row_num, worksheet)

    workbook.save(response)
    return response


def _xls_reviews_report(request, condominium, reviews, initial, until):
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml')
    response['Content-Disposition'] = 'attachment; filename="Relatorio_Avaliacoes.xlsx"'

    workbook = Workbook()
    worksheet = workbook.active

    worksheet.merge_cells('A1:D1')
    worksheet.merge_cells('A2:D2')

    _add_xls_header(condominium, worksheet, "Relatório de Avaliações", initial, until)

    columns = ['Serviço', 'Nota média', 'Prestador', 'Nº de Respostas']

    row_num = 3

    current_site = get_current_site(request)

    _add_xls_columns(columns, row_num, worksheet)

    for review in reviews:
        row_num += 1
        answers = ReviewAnswer.objects.filter(review=review, is_valid=True)

        row = [review.service, answers.aggregate(avg_score=Avg('rate'))['avg_score'], review.provider, answers.count()]

        _add_cell_values(row, row_num, worksheet)

    workbook.save(response)
    return response


def _xls_surveys_report(request, condominium, surveys, initial, until):
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml')
    response['Content-Disposition'] = 'attachment; filename="Relatorio_Enquetes.xlsx"'

    workbook = Workbook()
    worksheet = workbook.active

    worksheet.merge_cells('A1:E1')
    worksheet.merge_cells('A2:E2')

    _add_xls_header(condominium, worksheet, "Relatório de Enquetes", initial, until)

    columns = ['Pergunta', 'Respostas possíveis', 'Votos por resposta', 'Total de votos', 'Localização']

    row_num = 3

    current_site = get_current_site(request)

    _add_xls_columns(columns, row_num, worksheet)

    for survey in surveys:
        row_num += 1
        options = SurveyOptionModel.objects.filter(survey=survey)
        option_list = []
        votes_list = []
        total_votes = options.aggregate(Sum('votes'))['votes__sum'] or 0

        for option in options:
            option_list.append(option.option)
            votes_list.append(option.votes)

        options_values = ' | '.join(str(op) for op in option_list)
        options_votes = ' | '.join(str(vt) for vt in votes_list)

        row = [survey.question, options_values, options_votes, total_votes, survey.location.address if survey.location else ""]

        _add_cell_values(row, row_num, worksheet)

    workbook.save(response)
    return response


def _xls_contracts_report(request, condominium, contracts, initial, until):
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml')
    response['Content-Disposition'] = 'attachment; filename="Relatorio_Vencimentos.xlsx"'

    workbook = Workbook()
    worksheet = workbook.active

    worksheet.merge_cells('A1:G1')
    worksheet.merge_cells('A2:G2')

    _add_xls_header(condominium, worksheet, "Relatório de Vencimentos", initial, until)

    columns = ['Nome', 'Descrição', 'Ultima manutenção', 'Próxima manutenção', 'Email notificado',
               'Data da notificação', 'Imagens']

    row_num = 3

    current_site = get_current_site(request)

    _add_xls_columns(columns, row_num, worksheet)

    for contract in contracts:
        row_num += 1

        image = ""

        if contract.image:
            image = "https://" + current_site.domain + contract.image.url

        row = [contract.item, contract.description if contract.description else "",
               contract.last_maintenance.strftime('%d/%m/%Y'), contract.next_maintenance.strftime('%d/%m/%Y'),
               contract.to_email, contract.notify_day.strftime('%d/%m/%Y'), image]

        _add_cell_values(row, row_num, worksheet)

    workbook.save(response)
    return response


def _xls_checklists_report(request, condominium, checklists, initial, until):
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml')
    response['Content-Disposition'] = 'attachment; filename="Relatorio_Checklists.xlsx"'

    workbook = Workbook()
    worksheet = workbook.active

    worksheet.merge_cells('A1:E1')
    worksheet.merge_cells('A2:E2')

    _add_xls_header(condominium, worksheet, "Relatório de Checklists", initial, until)

    columns = ['Nome', 'Criado em', 'Verificações', 'Situação', 'Localização']

    row_num = 3

    _add_xls_columns(columns, row_num, worksheet)

    for checklist in checklists:
        row_num += 1

        checklist_tasks = Task.objects.filter(checklist=checklist)
        tasks_list = []
        status_list = []

        for task in checklist_tasks:
            tasks_list.append(task.task_name)

            if task.is_completed:
                status_list.append("OK")
            elif task.reported_problem:
                status_list.append(task.problem_description)
            else:
                status_list.append("Falta verificação")

        tasks_values = ' | '.join(str(tk) for tk in tasks_list)
        status_values = ' | '.join(str(st) for st in status_list)

        row = [checklist.title, checklist.created.astimezone(FIXED_TZ).strftime('%d/%m/%Y'), tasks_values, status_values
               , checklist.location.address if checklist.location else ""]

        _add_cell_values(row, row_num, worksheet)

    workbook.save(response)
    return response


def _xls_locations_report(request, condominium, locations, initial, until):
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml')
    response['Content-Disposition'] = 'attachment; filename="Relatorio_Localizacoes.xlsx"'

    workbook = Workbook()
    worksheet = workbook.active

    worksheet.merge_cells('A1:F1')
    worksheet.merge_cells('A2:F2')

    _add_xls_header(condominium, worksheet, "Relatório de Localizações", initial, until)

    columns = ['Usuário', 'IP', 'Endereço', 'Latitude', 'Longitude', 'Data']

    row_num = 3

    _add_xls_columns(columns, row_num, worksheet)

    for location in locations:
        row_num += 1

        row = [location.condominium.condominium_name, location.ip_address, location.address, location.latitude,
               location.longitude, location.created.astimezone(FIXED_TZ).strftime('%d/%m/%Y')]

        _add_cell_values(row, row_num, worksheet)

    workbook.save(response)
    return response


def _xls_visitants_report(request, condominium, visitants, initial, until):
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml')
    response['Content-Disposition'] = 'attachment; filename="Relatorio_Visitantes.xlsx"'

    workbook = Workbook()
    worksheet = workbook.active

    worksheet.merge_cells('A1:L1')
    worksheet.merge_cells('A2:L2')

    _add_xls_header(condominium, worksheet, "Relatório de Visitantes", initial, until)

    columns = ['Nome', 'Bloco', 'Apartamento', 'Liberado por', 'Liberado até', 'Porteiro que liberou', 'Chegada',
               'Saída', 'Documento', 'Veículo', 'Placa', 'Foto']

    row_num = 3

    _add_xls_columns(columns, row_num, worksheet)

    current_site = get_current_site(request)

    for visitant in visitants:
        row_num += 1

        visit_in = visitant.visit_in.astimezone(FIXED_TZ).strftime('%d/%m/%Y') if visitant.visit_in else ""
        leaves_in = visitant.leaves_in.astimezone(FIXED_TZ).strftime('%d/%m/%Y') if visitant.leaves_in else ""
        resident_name = visitant.resident.condominium_name if visitant.resident else ""
        until_value = visitant.until.strftime('%d/%m/%Y') if visitant.until else ""

        image = ""
        if visitant.photo:
            image = "https://" + current_site.domain + visitant.photo.url

        row = [visitant.name, visitant.block, visitant.apartment, resident_name,
               until_value, visitant.security_name, visit_in, leaves_in, visitant.document,
               visitant.vehicle_model if visitant.vehicle_model else "",
               visitant.vehicle_plate if visitant.vehicle_plate else "", image]

        _add_cell_values(row, row_num, worksheet)

    workbook.save(response)
    return response


def _xls_timelines_report(request, condominium, timelines, initial, until):
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml')
    response['Content-Disposition'] = 'attachment; filename="Relatorio_Timelines.xlsx"'

    workbook = Workbook()
    worksheet = workbook.active

    worksheet.merge_cells('A1:E1')
    worksheet.merge_cells('A2:E2')

    _add_xls_header(condominium, worksheet, "Relatório de Timelines", initial, until)

    columns = ['Título', 'Descrição', 'Início', 'Fim', 'Fases']

    row_num = 3

    _add_xls_columns(columns, row_num, worksheet)

    current_site = get_current_site(request)

    for timeline in timelines:
        row_num += 1

        phases = TimelinePhase.objects.filter(timeline=timeline).order_by('-end_date')

        phases_list = []
        for phase in phases:
            phases_list.append(phase.title)

        phases_values = ' | '.join(str(tk) for tk in phases_list)

        row = [timeline.title, timeline.description if timeline.description else "",
               timeline.start_date.strftime('%d/%m/%Y'), timeline.end_date.strftime('%d/%m/%Y'),
               phases_values]

        _add_cell_values(row, row_num, worksheet)

    workbook.save(response)
    return response


def _xls_resident_activities_report(request, condominium, resident_activities, initial, until):
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml')
    response['Content-Disposition'] = 'attachment; filename="Relatorio_Solicitacoes_Moradores.xlsx"'

    workbook = Workbook()
    worksheet = workbook.active

    worksheet.merge_cells('A1:I1')
    worksheet.merge_cells('A2:I2')

    _add_xls_header(condominium, worksheet, "Relatório de Solicitações", initial, until)

    columns = ['Título', 'Tipo', 'Descrição', 'Morador', 'Apartamento', 'Protocolo', 'Situação', 'Data',
               'Nº de Respostas']

    row_num = 3

    _add_xls_columns(columns, row_num, worksheet)

    for activity in resident_activities:
        row_num += 1

        answers = ResidentActivityAnswer.objects.filter(activity=activity).order_by('-created')
        apartment = str(activity.apartment.number) + " " + activity.apartment.complement if activity.apartment else ""

        row = [activity.title, activity.kind , activity.description, activity.resident, apartment, activity.protocol,
               activity.status, activity.created.astimezone(FIXED_TZ).strftime('%d/%m/%Y'), answers.count()]

        _add_cell_values(row, row_num, worksheet)

    workbook.save(response)
    return response


def _xls_booking_report(request, condominium, reservations, initial, until):
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml')
    response['Content-Disposition'] = 'attachment; filename="Relatorio_Reservas.xlsx"'

    workbook = Workbook()
    worksheet = workbook.active

    worksheet.merge_cells('A1:H1')
    worksheet.merge_cells('A2:H2')

    _add_xls_header(condominium, worksheet, "Relatório de Reservas", initial, until)

    columns = ['Área', 'Data', 'Horário', 'Morador', 'Situação', 'Aprovado por', 'Cancelado por', 'Realizada em']

    row_num = 3

    _add_xls_columns(columns, row_num, worksheet)

    for booking in reservations:
        row_num += 1
        time = booking.time.init_time.strftime('%H:%M') + ' - ' + booking.time.end_time.strftime('%H:%M')
        row = [booking.place.name, booking.date.strftime('%d/%m/%Y'), time, booking.resident.condominium_name,
               booking.status, booking.approved_by if booking.approved_by else "",
               booking.canceled_by if booking.canceled_by else "", booking.created.astimezone(FIXED_TZ).strftime('%d/%m/%Y')]

        _add_cell_values(row, row_num, worksheet)

    workbook.save(response)
    return response


def _xls_sessions_report(request, condominium, sessions, initial, until):
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml')
    response['Content-Disposition'] = 'attachment; filename="Relatorio_Sessoes.xlsx"'

    workbook = Workbook()
    worksheet = workbook.active

    worksheet.merge_cells('A1:G1')
    worksheet.merge_cells('A2:G2')

    _add_xls_header(condominium, worksheet, "Relatório de Sessões", initial, until)

    columns = ['Usuário', 'IP', 'Endereço', 'Data', 'Check-in', 'Check-out', 'Duração']

    row_num = 3

    _add_xls_columns(columns, row_num, worksheet)

    for control in sessions:
        row_num += 1

        duration = "Sessão ativa"

        if control.check_out:
            if control.session_time:
                days, seconds = divmod(control.session_time.total_seconds(), 86400)
                hours, remainder = divmod(seconds, 3600)
                minutes, seconds = divmod(remainder, 60)
                if days:

                    dur_string = f"{int(days)} days, {int(hours):02}:{int(minutes):02}:{int(seconds):02}"
                else:
                    dur_string = f"{int(hours):02}:{int(minutes):02}:{int(seconds):02}"

                duration = dur_string

        row = [control.user.condominium_name, control.ip_address, control.address, control.created.astimezone(FIXED_TZ).strftime('%d/%m/%Y'),
               control.check_in.strftime('%H:%M'), control.check_out.strftime('%H:%M') if control.check_out else "",
               duration]

        _add_cell_values(row, row_num, worksheet)

    workbook.save(response)
    return response


def _xls_storage_report(request, condominium, products, initial, until):
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml')
    response['Content-Disposition'] = 'attachment; filename="Relatorio_Estoque.xlsx"'

    workbook = Workbook()
    worksheet = workbook.active

    worksheet.merge_cells('A1:E1')
    worksheet.merge_cells('A2:E2')

    _add_xls_header(condominium, worksheet, "Relatório de Estoque", initial, until)

    columns = ['Nome', 'Descrição', 'Imagem', 'Quantidade', 'Quantidade mínima']

    row_num = 3

    _add_xls_columns(columns, row_num, worksheet)

    current_site = get_current_site(request)

    for product in products:
        row_num += 1

        image = "https://" + current_site.domain + product.image.url if product.image else ""
        row = [product.name, product.description, image, product.quantity, product.warning_quantity]

        _add_cell_values(row, row_num, worksheet)

    workbook.save(response)
    return response


def _add_xls_header(condominium, worksheet, report_name, initial = None, until = None):
    first_cell = worksheet['A1']
    first_cell.value = condominium.condominium_name
    first_cell.fill = PatternFill('solid', fgColor="1c4587")
    first_cell.font = Font(bold=True, color="ffffff")
    first_cell.alignment = Alignment(horizontal="center", vertical="center")
    second_cell = worksheet['A2']
    if initial and until:
        second_cell.value = report_name + " de " + parse_date(initial).strftime("%d/%m/%y") + " até " + parse_date(until).strftime("%d/%m/%y")
    else:
        second_cell.value = report_name
    second_cell.font = Font(bold=True, color="000000")
    second_cell.alignment = Alignment(horizontal="center", vertical="center")
    worksheet.title = report_name


def _add_xls_columns(columns, row_num, worksheet):
    for col_num, col_title in enumerate(columns, 1):
        cell = worksheet.cell(row=row_num, column=col_num)
        cell.value = col_title
        cell.fill = PatternFill('solid', fgColor="1c4587")
        cell.font = Font(bold=True, color="ffffff")
        cell.alignment = Alignment(horizontal="center", vertical="center")


def _add_cell_values(row, row_num, worksheet):
    for col_num, cell_value in enumerate(row, 1):
        cell = worksheet.cell(row=row_num, column=col_num)
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.value = cell_value


class PDF(FPDF):

    def __init__(self, test=None, *args, **kwargs):
        # Initialize parent class
        super().__init__(*args, **kwargs)
        # Store the watermark text if provided
        self.test = test

    def header(self):
        self.set_font("helvetica", "", 8)
        self.cell(0, 7, txt="AppGroup - AppPortaria", ln=True, align="R")
        self.set_text_color(255, 0, 0)
        self.cell(0, 7, txt="Para buscar use o atalho CTRL + F", ln=True, align="R")
        self.set_text_color(0, 0, 0)
        if self.test:
            _add_pdf_watermark(self)

    def footer(self):
        self.set_y(-15)
        self.set_font("helvetica", "I", 8)
        self.cell(0, 7, txt=f"Página {self.page_no()}/{{nb}}", align="C")



def aux_report(request):
    condominium = get_condominium(request)
    objects = VisitantReport.objects.all().order_by("-created")[:100] # Modify as per your model

    pdf = PDF("P", "mm" )
    pdf.add_page()
    pdf.set_auto_page_break(auto=True, margin=15)

    _add_pdf_header(pdf, condominium)

    pdf.set_font("helvetica", "B", 12)
    pdf.cell(0, 10, txt="Relatório dos Visitantes de 01/01/2025 até 10/01/2025", align="C", ln=True)

    pdf.set_font("helvetica", "", 10)
    pdf.cell(145, 7, txt="Visitas:", ln=False)
    pdf.cell(50, 7, txt=f"Total de visitantes: {len(objects)}", ln=True)

    bullet = "-"
    for idx, obj in enumerate(objects, start=1):
        pdf.cell(15)
        pdf.cell(200, 7, txt=f"{idx}. Visitante: {obj.name}", ln=True)
        y_start = pdf.get_y()
        pdf.cell(25)
        pdf.cell(200, 7, txt=f"{bullet} Bloco: {obj.block}", ln=True)
        pdf.cell(25)
        pdf.cell(200, 7, txt=f"{bullet} Apartamento: {obj.apartment}", ln=True)
        pdf.cell(25)
        pdf.cell(200, 7, txt=f"{bullet} Morador: {obj.resident.condominium_name}", ln=True)
        if obj.until:
            pdf.cell(25)
            pdf.cell(200, 7, txt=f"{bullet} Liberado até: {obj.until.strftime('%H:%M de %d/%m/%Y')}", ln=True)
        pdf.cell(25)
        pdf.cell(200, 7, txt=f"{bullet} Porteiro que liberou: {obj.security_name}", ln=True)
        if obj.visit_in:
            pdf.cell(25)
            pdf.cell(200, 7, txt=f"{bullet} Chegada: {obj.visit_in.astimezone(FIXED_TZ).strftime('%H:%M de %d/%m/%Y')}", ln=True)
        if obj.leaves_in:
            pdf.cell(25)
            pdf.cell(200, 7, txt=f"{bullet} Saída: {obj.leaves_in.strftime('%H:%M de %d/%m/%Y')}", ln=True)
        pdf.cell(25)
        pdf.cell(200, 7, txt=f"{bullet} Documento: {obj.document}", ln=True)
        pdf.cell(25)
        pdf.cell(200, 7, txt=f"{bullet} Modelo do veículo: {obj.vehicle_model}", ln=True)
        pdf.cell(25)
        pdf.cell(200, 7, txt=f"{bullet} Placa do veículo: {obj.vehicle_plate}", ln=True)
        if obj.photo:
            file_path = obj.photo.path
            if os.path.exists(file_path):
                pdf.image(obj.photo.path, pdf.w - 115, y_start, 100)
        pdf.ln(7)
        line_start = pdf.get_y()
        pdf.line(25, line_start, 185, line_start)
        pdf.ln(15)

    return _make_pdf_response(pdf, "Relatório_Visitantes")

def _add_pdf_header(pdf, condominium):
    try:
        logo = ReportLogo.objects.get(condominium=condominium)
        file_path = logo.image.path
        if os.path.exists(file_path):
            pdf.image(logo.image.path, 10, 8, 30)
    except ReportLogo.DoesNotExist:
        pass

    pdf.set_font("helvetica", "B", 14)
    pdf.cell(40)
    pdf.cell(200, 7, txt=f"Condomínio: {condominium.condominium_name}", ln=True)
    pdf.set_font("helvetica", "", 10)
    pdf.cell(40)
    if len(condominium.address) > 70:
        address_aux = condominium.address[:70]
        space_pos = address_aux.rfind(' ', 0, 69)
        pdf.multi_cell(200, 7, txt=f"Endereço: {condominium.address[:space_pos]}\n{condominium.address[space_pos:]}", ln=True)
    else:
        pdf.cell(200, 7, txt=f"Endereço: {condominium.address}", ln=True)
    pdf.cell(40)
    pdf.cell(200, 7, txt=f"Síndico/Responsável: {condominium.liquidator_name}", ln=True)
    if condominium.admin_name:
        pdf.cell(40)
        pdf.cell(200, 7, txt=f"Administradora: {condominium.admin_name}", ln=True)
    pdf.cell(40)
    pdf.cell(200, 7, txt=f"Whatsapp: {condominium.whatsapp}", ln=True)
    pdf.cell(40)
    pdf.cell(200, 7, txt=f"Email: {condominium.email}", ln=True)
    if condominium.site:
        pdf.cell(40)
        pdf.cell(200, 7, txt=f"Site: {condominium.site}", ln=True)
    pdf.ln(20)

def _make_pdf_response(pdf, file_name):
    pdf_output = BytesIO()
    pdf.output(pdf_output)
    pdf_output.seek(0)
    response = HttpResponse(pdf_output.read(), content_type='application/pdf')
    response['Content-Disposition'] = f'attachment; filename="{file_name}.pdf"'

    return response


def _add_pdf_watermark(pdf):
    pdf.set_font('helvetica', 'B', 40)

    # Set color for the watermark (light gray for opacity effect)
    pdf.set_text_color(200, 200, 200)

    # Rotate and position the watermark
    pdf.rotate(45, 60, 150)  # Rotate 45 degrees, position at x=60, y=150
    pdf.text(60, 150, 'PERÍODO DE TESTE')  # Add the watermark text

    # Reset rotation
    pdf.rotate(0)


def pedestrian_report(request):
    condominium = get_condominium(request)
    form = VisitantReportForm()

    if request.method == "POST":
        visits_from = request.POST.get('visits_from')
        visits_until = request.POST.get('visits_until')
        block = request.POST.get('block')
        start_datetime, end_datetime = _build_report_datetime_range(visits_from, visits_until)

        request.session['initial'] = visits_from
        request.session['until'] = visits_until

        if block:
            objects = Pedestrian.objects.filter(created__gte=start_datetime, condominium=condominium,
                                                    created__lt=end_datetime, destination__icontains=block)
        else:
            objects = Pedestrian.objects.filter(created__gte=start_datetime, condominium=condominium,
                                                created__lt=end_datetime)

        type = request.POST.get('type')

        if type == '1':

            return _pedestrians_pdf(request, condominium, objects, visits_from, visits_until)
        else:
            return _xls_pedestrians_report(request, condominium, objects, visits_from, visits_until)

    context = {'form': form,
               }

    return render(request, "info/condominium/report/pedestrian_report.html", context=context)


def _pedestrians_pdf(request, condominium, visitants, visits_from, visits_until):
    if len(visitants) > 0:
        context = {'condominium': condominium,
                   'visitants': visitants,
                   'from': visits_from,
                   'until': visits_until,
                   'test': condominium.is_testing}

        pdf = PDF(test=condominium.is_testing)
        pdf.add_page()
        pdf.set_auto_page_break(auto=True, margin=15)

        _add_pdf_header(pdf, condominium)

        pdf.set_font("helvetica", "B", 12)
        pdf.cell(0, 10, txt=f"Relatório dos Pedestres de {datetime.strptime(visits_from, '%Y-%m-%d').strftime('%d/%m/%Y')} até {datetime.strptime(visits_until, '%Y-%m-%d').strftime('%d/%m/%Y')}", align="C", ln=True)

        pdf.set_font("helvetica", "", 10)
        pdf.cell(145, 7, txt="Pedestres Cadastrados:", ln=False)
        pdf.cell(50, 7, txt=f"Total de pedestres: {len(visitants)}", ln=True)

        bullet = "-"
        for idx, obj in enumerate(visitants, start=1):
            pdf.cell(15)
            pdf.cell(200, 7, txt=f"{idx}. Pedestre: {obj.name}", ln=True)
            pdf.cell(25)
            pdf.cell(200, 7, txt=f"{bullet} Data da entrada: {obj.created.astimezone(FIXED_TZ).strftime('%H:%M de %d/%m/%Y')}", ln=True)
            y_start = pdf.get_y()
            pdf.cell(25)
            pdf.cell(200, 7, txt=f"{bullet} Protocolo: {obj.protocol}", ln=True)
            pdf.cell(25)
            pdf.cell(200, 7, txt=f"{bullet} Document: {obj.document}", ln=True)
            if obj.document_file:
                pdf.cell(15)
                pdf.set_text_color(0, 0, 255)  # Set text color to blue (optional)
                pdf.set_font(style='U')  # Underline the text (optional)
                pdf.cell(0, 10, f"{get_current_site(request).domain + obj.document_file.url}", ln=True,
                         link="http://" + get_current_site(request).domain + obj.document_file.url)
            pdf.set_text_color(0, 0, 0)
            pdf.set_font("helvetica", "", 10)
            pdf.cell(25)
            pdf.cell(200, 7, txt=f"{bullet} Destino: {obj.destination}", ln=True)
            pdf.cell(25)
            pdf.cell(200, 7, txt=f"{bullet} Observação: {obj.obs}", ln=True)
            pdf.cell(25)
            pdf.cell(200, 7, txt=f"{bullet} Autorizado por: {obj.authorized_by}", ln=True)

            if obj.has_leaved:
                pdf.cell(25)
                pdf.cell(200, 7, txt=f"{bullet} Saída: {obj.leaved_in.strftime('%H:%M de %d/%m/%Y')}", ln=True)
            if obj.photo:
                file_path = obj.photo.path
                if os.path.exists(file_path):
                    pdf.image(obj.photo.path, pdf.w - 65, y_start, 50)
            pdf.ln(7)
            line_start = pdf.get_y()
            pdf.line(25, line_start, 185, line_start)
            pdf.ln(15)

            messages.success(request, "Relatório Gerado!")
        return _make_pdf_response(pdf, "Relatório_Pedestres")

    else:
        messages.error(request, "Nenhum pedestre liberado neste período!")
    return redirect(reverse('info:pedestrian_report'))


def _xls_pedestrians_report(request, condominium, visitants, initial, until):
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml')
    response['Content-Disposition'] = 'attachment; filename="Relatorio_Pedestres.xlsx"'

    workbook = Workbook()
    worksheet = workbook.active

    worksheet.merge_cells('A1:J1')
    worksheet.merge_cells('A2:J2')

    _add_xls_header(condominium, worksheet, "Relatório de Pedestres", initial, until)

    columns = ['Protocolo', 'Nome', 'Documento', 'Arquivo', 'Destino', 'Observação', 'Autorizado por', 'Chegada',
               'Saída', 'Foto']

    row_num = 3

    _add_xls_columns(columns, row_num, worksheet)

    current_site = get_current_site(request)

    for visitant in visitants:
        row_num += 1

        visit_in = visitant.created.astimezone(FIXED_TZ).strftime('%d/%m/%Y')
        leaved_in = visitant.leaved_in.astimezone(FIXED_TZ).strftime('%d/%m/%Y') if visitant.has_leaved and visitant.leaved_in else ""

        image = ""
        if visitant.photo:
            image = "https://" + current_site.domain + visitant.photo.url

        file = ""
        if visitant.document_file:
            file = "https://" + current_site.domain + visitant.document_file.url

        row = [visitant.protocol ,visitant.name, visitant.document, file, visitant.destination, visitant.obs,
               visitant.authorized_by, visit_in, leaved_in, image]

        _add_cell_values(row, row_num, worksheet)

    workbook.save(response)
    messages.success(request, "Relatório Gerado!")
    return response


def vehicle_report(request):
    condominium = get_condominium(request)
    form = VisitantReportForm()

    if request.method == "POST":
        visits_from = request.POST.get('visits_from')
        visits_until = request.POST.get('visits_until')
        block = request.POST.get('block')
        start_datetime, end_datetime = _build_report_datetime_range(visits_from, visits_until)

        request.session['initial'] = visits_from
        request.session['until'] = visits_until

        if block:
            objects = Vehicle.objects.filter(created__gte=start_datetime, condominium=condominium,
                                                    created__lt=end_datetime, destination__icontains=block)
        else:
            objects = Vehicle.objects.filter(created__gte=start_datetime, condominium=condominium,
                                                created__lt=end_datetime)

        type = request.POST.get('type')

        if type == '1':

            return _vehicles_pdf(request, condominium, objects, visits_from, visits_until)
        else:
            return _xls_vehicles_report(request, condominium, objects, visits_from, visits_until)

    context = {'form': form,
               }

    return render(request, "info/condominium/report/vehicle_report.html", context=context)


def _vehicles_pdf(request, condominium, visitants, visits_from, visits_until):
    if len(visitants) > 0:
        context = {'condominium': condominium,
                   'visitants': visitants,
                   'from': visits_from,
                   'until': visits_until,
                   'test': condominium.is_testing}

        pdf = PDF(test=condominium.is_testing)
        pdf.add_page()
        pdf.set_auto_page_break(auto=True, margin=15)

        _add_pdf_header(pdf, condominium)

        pdf.set_font("helvetica", "B", 12)
        pdf.cell(0, 10, txt=f"Relatório dos Veículos de {datetime.strptime(visits_from, '%Y-%m-%d').strftime('%d/%m/%Y')} até {datetime.strptime(visits_until, '%Y-%m-%d').strftime('%d/%m/%Y')}", align="C", ln=True)

        pdf.set_font("helvetica", "", 10)
        pdf.cell(145, 7, txt="Veículos Cadastrados:", ln=False)
        pdf.cell(50, 7, txt=f"Total de veículos: {len(visitants)}", ln=True)

        bullet = "-"
        for idx, obj in enumerate(visitants, start=1):
            pdf.cell(15)
            pdf.cell(200, 7, txt=f"{idx}. Nome: {obj.name}", ln=True)
            pdf.cell(25)
            pdf.cell(200, 7, txt=f"{bullet} Data da entrada: {obj.created.astimezone(FIXED_TZ).strftime('%H:%M de %d/%m/%Y')}", ln=True)
            y_start = pdf.get_y()
            pdf.cell(25)
            pdf.cell(200, 7, txt=f"{bullet} Protocolo: {obj.protocol}", ln=True)
            pdf.cell(25)
            pdf.cell(200, 7, txt=f"{bullet} Document: {obj.document}", ln=True)
            if obj.document_file:
                pdf.cell(15)
                pdf.set_text_color(0, 0, 255)  # Set text color to blue (optional)
                pdf.set_font(style='U')  # Underline the text (optional)
                pdf.cell(0, 10, f"{get_current_site(request).domain + obj.document_file.url}", ln=True,
                         link="http://" + get_current_site(request).domain + obj.document_file.url)
            pdf.set_text_color(0, 0, 0)
            pdf.set_font("helvetica", "", 10)
            pdf.cell(25)
            pdf.cell(200, 7, txt=f"{bullet} Veículo: {obj.vehicle}", ln=True)
            pdf.cell(25)
            pdf.cell(200, 7, txt=f"{bullet} Placa: {obj.vehicle_plate}", ln=True)
            pdf.cell(25)
            pdf.cell(200, 7, txt=f"{bullet} Destino: {obj.destination}", ln=True)
            pdf.cell(25)
            pdf.cell(200, 7, txt=f"{bullet} Observação: {obj.obs}", ln=True)
            pdf.cell(25)
            pdf.cell(200, 7, txt=f"{bullet} Autorizado por: {obj.authorized_by}", ln=True)

            if obj.has_leaved:
                pdf.cell(25)
                pdf.cell(200, 7, txt=f"{bullet} Saída: {obj.leaved_in.strftime('%H:%M de %d/%m/%Y')}", ln=True)
            if obj.photo:
                file_path = obj.photo.path
                if os.path.exists(file_path):
                    pdf.image(obj.photo.path, pdf.w - 65, y_start, 50)
            pdf.ln(7)
            line_start = pdf.get_y()
            pdf.line(25, line_start, 185, line_start)
            pdf.ln(15)

            messages.success(request, "Relatório Gerado!")
        return _make_pdf_response(pdf, "Relatório_Veículos")

    else:
        messages.error(request, "Nenhum veículo liberado neste período!")
    return redirect(reverse('info:vehicle_report'))


def _xls_vehicles_report(request, condominium, visitants, initial, until):
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml')
    response['Content-Disposition'] = 'attachment; filename="Relatorio_Veiculos.xlsx"'

    workbook = Workbook()
    worksheet = workbook.active

    worksheet.merge_cells('A1:L1')
    worksheet.merge_cells('A2:L2')

    _add_xls_header(condominium, worksheet, "Relatório de Veículos", initial, until)

    columns = ['Protocolo', 'Nome', 'Documento', 'Arquivo', 'Veículo', 'Placa', 'Destino', 'Observação', 'Autorizado por', 'Chegada',
               'Saída', 'Foto']

    row_num = 3

    _add_xls_columns(columns, row_num, worksheet)

    current_site = get_current_site(request)

    for visitant in visitants:
        row_num += 1

        visit_in = visitant.created.astimezone(FIXED_TZ).strftime('%d/%m/%Y')
        leaved_in = visitant.leaved_in.astimezone(FIXED_TZ).strftime('%d/%m/%Y') if visitant.has_leaved and visitant.leaved_in else ""

        image = ""
        if visitant.photo:
            image = "https://" + current_site.domain + visitant.photo.url

        file = ""
        if visitant.document_file:
            file = "https://" + current_site.domain + visitant.document_file.url

        row = [visitant.protocol ,visitant.name, visitant.document, file, visitant.vehicle, visitant.vehicle_plate,
               visitant.destination, visitant.obs, visitant.authorized_by, visit_in, leaved_in, image]

        _add_cell_values(row, row_num, worksheet)

    workbook.save(response)
    messages.success(request, "Relatório Gerado!")
    return response
